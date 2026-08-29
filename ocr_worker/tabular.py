"""Clean already-structured Excel/CSV/XLS workbooks with header-row detection."""
from __future__ import annotations

import hashlib
import sqlite3
from pathlib import Path
from typing import Any

from clean import canonical_header, clean_text, find_header_row, validate_detailed
from common import FileResult, YELLOW
from export import output_file


def clean_tabular(source: Path, master: dict[str, list[str]], output_dir: Path) -> FileResult:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    destination = output_file(source, output_dir)
    output_book = Workbook()
    output_book.remove(output_book.active)
    yellow = PatternFill(fill_type="solid", fgColor=YELLOW)
    duplicate_db = output_dir / f".{hashlib.sha1(str(source).encode()).hexdigest()}.sqlite"
    connection = sqlite3.connect(duplicate_db)
    connection.execute("CREATE TABLE seen (key TEXT PRIMARY KEY)")
    records = low_confidence = 0
    review_items: list[dict[str, Any]] = []
    warnings: list[str] = []
    try:
        sheets = list(_load_sheets(source))
        for title, rows in sheets:
            if not rows:
                continue
            header_at = find_header_row(rows)
            if header_at > 0:
                warnings.append(f"Skipped {header_at} row(s) above the header in '{title}'.")
            raw_headers = [str(value or f"column_{index + 1}").strip() or f"column_{index + 1}" for index, value in enumerate(rows[header_at])]
            fields = [canonical_header(value) for value in raw_headers]
            target = output_book.create_sheet(title=str(title)[:31])
            target.append(raw_headers)
            target.freeze_panes = "A2"
            for raw_row in rows[header_at + 1 :]:
                cleaned: list[str] = []
                changed = False
                for index, value in enumerate(raw_row):
                    field = fields[index] if index < len(fields) else "value"
                    result = validate_detailed(clean_text(value, field), field, master)
                    cleaned.append(result.value)
                    changed = changed or result.review
                if not any(cleaned):
                    continue
                key = hashlib.sha256("\x1f".join(cleaned).encode("utf-8")).hexdigest()
                try:
                    connection.execute("INSERT INTO seen VALUES (?)", (key,))
                except sqlite3.IntegrityError:
                    continue
                target.append(cleaned)
                records += 1
                if changed:
                    low_confidence += 1
                    for cell in target[target.max_row]:
                        cell.fill = yellow
                    review_items.append({
                        "output": str(destination),
                        "sheet": target.title,
                        "row": target.max_row,
                        "column": 1,
                        "header": raw_headers[0] if raw_headers else "",
                        "value": cleaned[0] if cleaned else "",
                        "confidence": "",
                        "suggestion": "master-data",
                    })
            if raw_headers:
                target.auto_filter.ref = target.dimensions
            connection.commit()
        if not output_book.worksheets:
            sheet = output_book.create_sheet("Cleaned Data")
            sheet.append(["Message"])
            sheet.append(["This file holds no rows that could be cleaned."])
        output_book.save(destination)
    finally:
        connection.close()
        try:
            duplicate_db.unlink(missing_ok=True)
        except OSError:
            pass
    if low_confidence:
        warnings.append("Some values were corrected against the local reference lists; check the highlighted rows.")
    return FileResult(
        str(source), str(destination), records=records, low_confidence=low_confidence,
        warnings=warnings or None, review_items=review_items,
        template={"source": source.name, "type": "tabular", "columns": []},
    )


def _load_sheets(source: Path) -> list[tuple[str, list[list[Any]]]]:
    suffix = source.suffix.lower()
    if suffix == ".csv":
        import csv
        with source.open("r", encoding="utf-8-sig", newline="") as stream:
            return [("Cleaned Data", [list(row) for row in csv.reader(stream)])]
    if suffix == ".xls":
        import xlrd
        book = xlrd.open_workbook(str(source))
        return [
            (sheet.name, [sheet.row_values(index) for index in range(sheet.nrows)])
            for sheet in book.sheets()
        ]
    from openpyxl import load_workbook
    book = load_workbook(source, read_only=True, data_only=True)
    try:
        sheets = []
        for worksheet in book.worksheets:
            sheets.append((worksheet.title, [list(row) for row in worksheet.iter_rows(values_only=True)]))
        return sheets
    finally:
        book.close()
