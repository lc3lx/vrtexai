"""Excel export, yellow review highlighting, status report, and review queue."""
from __future__ import annotations

import hashlib
import json
import re
import time
from pathlib import Path
from typing import Any, Sequence

from common import CONFIDENCE_THRESHOLD, FileResult, RED, YELLOW
from templates import build_template


def output_file(source: Path, output_dir: Path, suffix: str = ".xlsx") -> Path:
    digest = hashlib.sha1(str(source.resolve()).encode("utf-8")).hexdigest()[:8]
    return output_dir / f"{source.stem}_cleaned_{digest}{suffix}"


def _save_workbook(book: Any, destination: Path) -> Path:
    """Write the workbook. If the target is locked (open in Excel), use a new name."""
    try:
        book.save(destination)
        return destination
    except PermissionError:
        stamp = time.strftime("%Y%m%d_%H%M%S")
        alternate = destination.with_name(f"{destination.stem}_{stamp}{destination.suffix}")
        book.save(alternate)
        return alternate

def _paint(sheet, row_index: int, yellow, columns: Sequence[int] | None = None) -> None:
    cells = sheet[row_index] if columns is None else [sheet.cell(row_index, column) for column in columns]
    for cell in cells:
        cell.fill = yellow


def write_generic_tables(
    destination: Path,
    source: Path,
    tables: Sequence[dict[str, Any]],
) -> tuple[int, int, list[dict[str, Any]], dict[str, Any], Path]:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    book = Workbook()
    book.remove(book.active)
    yellow = PatternFill(fill_type="solid", fgColor=YELLOW)
    records = low = 0
    review_items: list[dict[str, Any]] = []
    first_headers: list[str] = []
    for table_index, table in enumerate(tables, start=1):
        headers: list[str] = list(table.get("headers") or [])
        rows: list[dict[str, Any]] = list(table.get("rows") or [])
        if not headers and rows:
            width = max(len(row.get("values") or []) for row in rows)
            headers = [f"column_{index + 1}" for index in range(width)]
        if not first_headers:
            first_headers = headers
        sheet = book.create_sheet(title=str(table.get("name") or f"Table{table_index}")[:31])
        sheet.append(headers)
        sheet.freeze_panes = "A2"
        for row in rows:
            values = list(row.get("values") or [])
            while len(values) < len(headers):
                values.append("")
            sheet.append(values[: len(headers)] if headers else values)
            records += 1
            excel_row = sheet.max_row
            confidences = list(row.get("confidences") or [])
            review_columns = {
                int(column)
                for column in row.get("review_columns") or []
                if isinstance(column, int) or str(column).isdigit()
            }
            review = bool(row.get("review"))
            if not review and confidences:
                review_columns.update(
                    index
                    for index, (value, conf) in enumerate(zip(values, confidences))
                    if str(value).strip() and conf is not None and conf < CONFIDENCE_THRESHOLD
                )
                review = bool(review_columns)
            if review:
                low += 1
                if not review_columns:
                    review_columns.add(next((index for index, value in enumerate(values) if str(value).strip()), 0))
                review_columns = {index for index in review_columns if 0 <= index < len(headers)}
                _paint(sheet, excel_row, yellow, [index + 1 for index in sorted(review_columns)])
                for index in sorted(review_columns):
                    value = values[index] if index < len(values) else ""
                    header = headers[index] if index < len(headers) else ""
                    conf = confidences[index] if index < len(confidences) else row.get("confidence")
                    review_items.append({
                        "output": str(destination),
                        "sheet": sheet.title,
                        "row": excel_row,
                        "column": index + 1,
                        "header": header,
                        "value": value,
                        "confidence": conf if conf is not None else "",
                        "suggestion": row.get("suggestion") or "",
                    })
        if headers:
            sheet.auto_filter.ref = sheet.dimensions
    if not book.worksheets:
        sheet = book.create_sheet(title="Extracted")
        sheet.append(["Message"])
        sheet.append(["No rows could be extracted from this file."])
    destination = _save_workbook(book, destination)
    for item in review_items:
        item["output"] = str(destination)
    template = build_template(source.name, "table", " ".join(first_headers), first_headers)
    return records, low, review_items, template, destination


def write_invoice(
    destination: Path,
    source: Path,
    parsed: dict[str, Any],
) -> tuple[int, int, list[dict[str, Any]], dict[str, Any], Path]:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    book = Workbook()
    yellow = PatternFill(fill_type="solid", fgColor=YELLOW)
    header_sheet = book.active
    header_sheet.title = "Header"
    header_sheet.append(["Field", "Value"])
    labels = [
        ("supplier", "Supplier"),
        ("client_name", "Client Name"),
        ("invoice_number", "Invoice Number"),
        ("invoice_date", "Invoice Date"),
        ("due_date", "Due Date"),
        ("tax_number", "Tax Number"),
        ("currency", "Currency"),
    ]
    header = parsed.get("header") or {}
    for key, label in labels:
        header_sheet.append([label, header.get(key, "")])
    header_sheet.freeze_panes = "A2"

    items_sheet = book.create_sheet("Items")
    item_headers = ["Description", "SKU", "Qty", "Unit Price", "Total", "Confidence", "Review Required"]
    items_sheet.append(item_headers)
    items_sheet.freeze_panes = "A2"
    review_items: list[dict[str, Any]] = []
    low = 0
    for item in parsed.get("items") or []:
        review_flag = "YES" if item.get("review") else ""
        items_sheet.append([
            item.get("description", ""), item.get("sku", ""), item.get("qty", ""), item.get("unit_price", ""),
            item.get("total", ""), item.get("confidence", ""), review_flag,
        ])
        if item.get("review"):
            low += 1
            _paint(items_sheet, items_sheet.max_row, yellow)
            review_items.append({
                "output": str(destination),
                "sheet": "Items",
                "row": items_sheet.max_row,
                "column": 1,
                "header": "Description",
                "value": item.get("description", ""),
                "confidence": item.get("confidence", ""),
                "suggestion": "",
            })
    if item_headers:
        items_sheet.auto_filter.ref = items_sheet.dimensions

    totals_sheet = book.create_sheet("Totals")
    totals_sheet.append(["Field", "Value"])
    totals = parsed.get("totals") or {}
    totals_sheet.append(["Subtotal", totals.get("subtotal") or header.get("subtotal", "")])
    totals_sheet.append(["Tax", totals.get("tax_amount") or header.get("tax_amount", "")])
    totals_sheet.append(["Grand Total", totals.get("grand_total") or header.get("grand_total", "")])
    totals_sheet.append(["Currency", totals.get("currency") or header.get("currency", "")])
    totals_sheet["B4"].font = Font(bold=True)
    destination = _save_workbook(book, destination)
    for item in review_items:
        item["output"] = str(destination)
    records = len(parsed.get("items") or [])
    template = build_template(source.name, "invoice", json.dumps(header, ensure_ascii=False), item_headers)
    return records, low, review_items, template, destination


def write_dynamic_workbook(
    destination: Path,
    source: Path,
    workbook: dict[str, Any],
) -> tuple[int, int, list[dict[str, Any]], dict[str, Any], Path]:
    """Write an LLM-designed workbook: sheet names and columns are not fixed."""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill

    book = Workbook()
    book.remove(book.active)
    yellow = PatternFill(fill_type="solid", fgColor=YELLOW)
    records = low = 0
    review_items: list[dict[str, Any]] = []
    first_headers: list[str] = []
    sheets = list(workbook.get("sheets") or [])
    if not sheets:
        sheets = [{
            "name": "Extracted",
            "columns": ["Message"],
            "rows": [["لم يُرجع النموذج أي جداول."]],
        }]

    for index, sheet_spec in enumerate(sheets, start=1):
        headers = [str(h).strip() or f"column_{i + 1}" for i, h in enumerate(sheet_spec.get("columns") or [])]
        raw_rows = list(sheet_spec.get("rows") or [])
        if not headers and raw_rows:
            width = max((len(r) if isinstance(r, (list, tuple)) else len(r.get("values") or []) for r in raw_rows), default=1)
            headers = [f"column_{i + 1}" for i in range(width)]
        if not first_headers:
            first_headers = headers
        title = str(sheet_spec.get("name") or f"Sheet{index}")[:31] or f"Sheet{index}"
        # Avoid duplicate sheet titles.
        base = title
        suffix = 2
        existing = {s.title for s in book.worksheets}
        while title in existing:
            title = f"{base[:28]}_{suffix}"[:31]
            suffix += 1
        sheet = book.create_sheet(title=title)
        if headers:
            sheet.append(headers)
            sheet.freeze_panes = "A2"
        for row in raw_rows:
            review = False
            confidences: list[float] = []
            if isinstance(row, dict):
                values = list(row.get("values") or [])
                review = bool(row.get("review"))
                confidences = [float(c) for c in (row.get("confidences") or []) if c is not None]
            else:
                values = list(row)
            while len(values) < len(headers):
                values.append("")
            if headers:
                values = values[: len(headers)]
            sheet.append(["" if v is None else v for v in values])
            records += 1
            excel_row = sheet.max_row
            if not review and confidences:
                review = any(
                    conf < CONFIDENCE_THRESHOLD
                    for conf, value in zip(confidences, values)
                    if str(value).strip()
                )
            # Mark sparse/empty critical cells for review.
            if not review and headers and not any(str(v).strip() for v in values):
                review = True
            if review:
                low += 1
                _paint(sheet, excel_row, yellow)
                review_items.append({
                    "output": str(destination),
                    "sheet": sheet.title,
                    "row": excel_row,
                    "column": 1,
                    "header": headers[0] if headers else "",
                    "value": values[0] if values else "",
                    "confidence": confidences[0] if confidences else "",
                    "suggestion": "",
                })
        if headers:
            sheet.auto_filter.ref = sheet.dimensions

    destination = _save_workbook(book, destination)
    for item in review_items:
        item["output"] = str(destination)
    template = build_template(
        source.name,
        "dynamic",
        " ".join(first_headers),
        first_headers,
    )
    return records, low, review_items, template, destination


def _sheet_title(source: Path, page: int, total_pages: int, taken: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", source.stem)[:26] or "Extracted"
    title = base if total_pages <= 1 else f"{base[:24]}-{page}"
    candidate, suffix = title[:31], 2
    while candidate in taken:
        candidate = f"{title[:28]}_{suffix}"[:31]
        suffix += 1
    taken.add(candidate)
    return candidate


def _needs_review(cell: dict[str, Any]) -> bool:
    if cell.get("review"):
        return True
    text = str(cell.get("text") or "").strip()
    confidence = cell.get("conf")
    return bool(text) and confidence is not None and float(confidence) < CONFIDENCE_THRESHOLD


def write_single_sheet(
    destination: Path,
    source: Path,
    documents: Sequence[dict[str, Any]],
) -> tuple[int, int, list[dict[str, Any]], dict[str, Any], Path]:
    """Write everything found in a document into one sheet per page.

    The customer asked for one page of Excel carrying *all* the data in the
    image, so sections are stacked vertically in reading order rather than
    split across a fixed set of sheets. Nothing that was read is dropped: text
    that belongs to no table still gets its own block.
    """
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    book = Workbook()
    book.remove(book.active)
    yellow = PatternFill(fill_type="solid", fgColor=YELLOW)
    title_font = Font(bold=True, size=12)
    header_font = Font(bold=True)
    records = low = 0
    review_items: list[dict[str, Any]] = []
    first_headers: list[str] = []
    taken: set[str] = set()

    for page_index, document in enumerate(documents, start=1):
        regions = list(document.get("regions") or [])
        sheet = book.create_sheet(title=_sheet_title(source, page_index, len(documents), taken))
        if str(document.get("direction") or "ltr") == "rtl":
            sheet.sheet_view.rightToLeft = True
        widths: dict[int, int] = {}
        row_number = 1

        for region in regions:
            rows = list(region.get("rows") or [])
            columns = list(region.get("columns") or [])
            if not rows and not columns:
                continue
            width = max([len(columns)] + [len(row) for row in rows]) or 1

            if region.get("title"):
                cell = sheet.cell(row_number, 1, str(region["title"]))
                cell.font = title_font
                if width > 1:
                    sheet.merge_cells(
                        start_row=row_number, start_column=1,
                        end_row=row_number, end_column=width,
                    )
                row_number += 1

            if columns:
                for column_index, name in enumerate(columns, start=1):
                    cell = sheet.cell(row_number, column_index, name)
                    cell.font = header_font
                    widths[column_index] = max(widths.get(column_index, 10), min(len(str(name)) + 2, 60))
                if not first_headers:
                    first_headers = [str(name) for name in columns]
                row_number += 1

            for row in rows:
                painted: list[int] = []
                for column_index, source_cell in enumerate(row, start=1):
                    text = str(source_cell.get("text") or "")
                    cell = sheet.cell(row_number, column_index, text)
                    cell.alignment = Alignment(vertical="top", wrap_text=len(text) > 60)
                    widths[column_index] = max(widths.get(column_index, 10), min(len(text) + 2, 60))
                    note = source_cell.get("note")
                    if note:
                        cell.comment = Comment(str(note), "Vertex")
                    if _needs_review(source_cell):
                        cell.fill = yellow
                        painted.append(column_index)
                        review_items.append({
                            "output": str(destination),
                            "sheet": sheet.title,
                            "row": row_number,
                            "column": column_index,
                            "header": columns[column_index - 1] if column_index <= len(columns) else "",
                            "value": text,
                            "confidence": source_cell.get("conf", ""),
                            # The competing reading of the same pixels, so the
                            # reviewer sees the alternative rather than guessing.
                            "suggestion": next(
                                (
                                    str(alternative.get("text") or "")
                                    for alternative in source_cell.get("alternatives") or []
                                    if str(alternative.get("text") or "").strip()
                                    and str(alternative.get("text")) != text
                                ),
                                "",
                            ),
                            "note": str(note or ""),
                        })
                if painted:
                    low += 1
                records += 1
                row_number += 1
            row_number += 1  # blank spacer between sections

        if row_number == 1:
            sheet.cell(1, 1, "لم يتم استخراج أي بيانات من هذه الصفحة.")
        for column_index, width_value in widths.items():
            sheet.column_dimensions[get_column_letter(column_index)].width = max(10, width_value)

    if not book.worksheets:
        sheet = book.create_sheet(title="Extracted")
        sheet.cell(1, 1, "لم يتم استخراج أي بيانات من هذا الملف.")

    destination = _save_workbook(book, destination)
    for item in review_items:
        item["output"] = str(destination)
    template = build_template(source.name, "document", " ".join(first_headers), first_headers)
    return records, low, review_items, template, destination


def write_status_report(output_dir: Path, results: list[FileResult]) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    destination = output_dir / "processing_status_report.xlsx"
    book = Workbook()
    sheet = book.active
    sheet.title = "Processing Status"
    sheet.append(["Source", "Status", "Output", "Records", "Low confidence", "Warnings / Error"])
    yellow, red = PatternFill(fill_type="solid", fgColor=YELLOW), PatternFill(fill_type="solid", fgColor=RED)
    for result in results:
        sheet.append([
            result.source, result.status_label, result.output or "", result.records,
            result.low_confidence, "; ".join(result.warnings or []) or result.error or "",
        ])
        fill = red if result.status in {"failed", "unsupported"} else yellow if result.low_confidence or result.warnings else None
        if fill:
            for cell in sheet[sheet.max_row]:
                cell.fill = fill
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    return _save_workbook(book, destination)

def write_review_queue(output_dir: Path, results: list[FileResult]) -> Path:
    destination = output_dir / "review_queue.json"
    items = []
    for result in results:
        items.extend(result.review_items or [])
    destination.write_text(json.dumps({"items": items}, ensure_ascii=False, indent=2), encoding="utf-8")
    return destination
