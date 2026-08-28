"""Field cleaning, date/number normalization, header detection, and master-data match."""
from __future__ import annotations

import csv
import datetime as dt
import re
from dataclasses import dataclass
from difflib import SequenceMatcher, get_close_matches
from pathlib import Path
from typing import Any, Iterable, Sequence

from common import DATE_FIELD, HEADER_ALIASES, ID_FIELD, NUMBER_FIELD, SPACE, SYMBOLS

ARABIC_DIGITS = str.maketrans("٠١٢٣٤٥٦٧٨٩۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")
DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%m/%d/%Y",
    "%d.%m.%Y", "%Y.%m.%d", "%d/%m/%y", "%d-%m-%y",
)


def clean_text(value: Any, field_name: str = "") -> str:
    text = SPACE.sub(" ", SYMBOLS.sub(" ", "" if value is None else str(value))).strip()
    text = text.translate(ARABIC_DIGITS)
    if ID_FIELD.search(field_name):
        text = correct_identifier(text, field_name)
    elif NUMBER_FIELD.search(field_name) and re.fullmatch(r"[0-9OoIl.,\s/+-]+", text or ""):
        text = text.replace("O", "0").replace("o", "0").replace("I", "1").replace("l", "1")
    return text


def correct_identifier(text: str, field_name: str = "") -> str:
    """Fix O/0 and I/l/1 confusion in numeric or code segments only."""
    if not text:
        return text
    if DATE_FIELD.search(field_name) and not ID_FIELD.search(field_name):
        return text

    def fix_part(part: str) -> str:
        if not part or part == "-":
            return part
        confusing = sum(ch in "OoIl" for ch in part)
        digits = sum(ch.isdigit() for ch in part)
        letters = sum(ch.isalpha() and ch not in "OoIl" for ch in part)
        if digits + confusing == 0:
            return part
        if letters and letters > digits + confusing:
            return part
        return (
            part.replace("O", "0")
            .replace("o", "0")
            .replace("I", "1")
            .replace("l", "1")
        )

    return "".join(fix_part(part) for part in re.split(r"(-)", text))


def normalize_date(value: str) -> str:
    value = value.strip().translate(ARABIC_DIGITS)
    if not value:
        return value
    for fmt in DATE_FORMATS:
        try:
            return dt.datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return value


def normalize_number(value: str) -> str:
    original = value.strip()
    text = original.translate(ARABIC_DIGITS).replace(" ", "")
    if not text:
        return original
    if re.fullmatch(r"\d{1,6}(,\d{3})+(\.\d+)?", text):
        return text.replace(",", "")
    if re.fullmatch(r"\d{1,6}(\.\d{3})+(,\d+)?", text):
        return text.replace(".", "").replace(",", ".")
    if text.count(".") > 1 and "," not in text and re.fullmatch(r"\d{1,6}(\.\d{3})+", text):
        return text.replace(".", "")
    return original.translate(ARABIC_DIGITS)


def canonical_header(value: str) -> str:
    def header_key(text: str) -> str:
        # Header labels are short and OCR commonly confuses Arabic alef forms
        # and the final taa marbuta.  Normalising those presentation variants
        # makes "الكمية" / "الكميه" equivalent without changing cell values.
        return (
            clean_text(text).casefold()
            .replace("أ", "ا")
            .replace("إ", "ا")
            .replace("آ", "ا")
            .replace("ى", "ي")
            .replace("ة", "ه")
            .replace("ـ", "")
        )

    normalized = header_key(value)
    for key, aliases in HEADER_ALIASES.items():
        normalized_aliases = [header_key(alias) for alias in aliases]
        if normalized in normalized_aliases or any(alias in normalized for alias in normalized_aliases):
            return key
    # A one-glyph error in a four-to-eight character Arabic label should not
    # make an otherwise valid invoice lose its schema.  This remains limited to
    # the explicit header aliases, never arbitrary data values.
    if len(normalized) >= 4:
        for key, aliases in HEADER_ALIASES.items():
            if any(
                len(alias_key) >= 4 and SequenceMatcher(None, normalized, alias_key).ratio() >= 0.78
                for alias_key in (header_key(alias) for alias in aliases)
            ):
                return key
    return clean_text(value) or "column"


def header_row_score(row: Sequence[Any]) -> float:
    texts = [str(cell or "").strip() for cell in row]
    filled = [text for text in texts if text]
    if len(filled) < 2:
        return -1.0
    if len(filled) <= 2 and any(":" in text or "：" in text for text in filled):
        return 0.2
    joined = " ".join(filled)
    joined_cf = joined.casefold()
    # Data rows with emails/phones/URLs are almost never headers.
    data_hits = 0
    if re.search(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", joined, re.I):
        data_hits += 2
    if re.search(r"https?://|www\.|instagram\.com", joined, re.I):
        data_hits += 2
    if re.search(r"\b(?:\+?966|05)\d{7,10}\b", joined):
        data_hits += 2
    if data_hits >= 2:
        return -2.0
    numeric = sum(1 for text in filled if re.fullmatch(r"[\d.,/\\-]+", text.translate(ARABIC_DIGITS)))
    if numeric / len(filled) > 0.45:
        return 0.1
    unique = len({text.casefold() for text in filled}) / len(filled)
    avg_len = sum(len(text) for text in filled) / len(filled)
    alias_hits = 0
    for aliases in HEADER_ALIASES.values():
        if any(alias in joined_cf for alias in aliases):
            alias_hits += 1
    length_penalty = 0.0 if avg_len <= 32 else min(avg_len / 80.0, 1.5)
    return len(filled) + unique * 2 + alias_hits * 3 - length_penalty - numeric - data_hits


def find_header_row(rows: Sequence[Sequence[Any]], max_scan: int = 40) -> int:
    best_index, best_score = 0, -1.0
    for index, row in enumerate(rows[:max_scan]):
        score = header_row_score(row)
        # Prefer earlier rows when scores are close so title noise below the
        # true header does not steal the header slot.
        if score > best_score + 0.35 or (abs(score - best_score) <= 0.35 and index < best_index):
            best_index, best_score = index, score
    if best_score < 1.0:
        return 0
    return best_index


def header_row_is_reliable(row: Sequence[Any]) -> bool:
    """Return whether an OCR row is credible enough to become Excel headers.

    OCR occasionally loses the real header row and promotes the first customer
    record instead.  Dropping a real record is worse than using neutral column
    names, so this deliberately has a conservative threshold.
    """
    texts = [str(cell or "").strip() for cell in row]
    filled = [text for text in texts if text]
    if len(filled) < 2:
        return False
    joined = " ".join(filled)
    if "@" in joined or re.search(r"https?://|www\.", joined, re.I):
        return False
    if any(len(text) > 96 for text in filled):
        return False
    average_length = sum(len(text) for text in filled) / len(filled)
    if average_length > 42:
        return False
    numeric_cells = sum(bool(re.search(r"\d", text.translate(ARABIC_DIGITS))) for text in filled)
    if numeric_cells > max(1, int(len(filled) * 0.35)):
        return False

    aliases = sum(
        1
        for candidates in HEADER_ALIASES.values()
        if any(candidate in joined.casefold() for candidate in candidates)
    )
    # Unknown but compact headings such as STATE / ST_ABB / YR are valid too.
    return aliases > 0 or (header_row_score(row) >= 4.0 and average_length <= 28)


def find_reliable_header_row(rows: Sequence[Sequence[Any]], max_scan: int = 40) -> int | None:
    """Find the strongest credible header row, or keep every OCR row as data."""
    best_index: int | None = None
    best_score = float("-inf")
    for index, row in enumerate(rows[:max_scan]):
        if not header_row_is_reliable(row):
            continue
        score = header_row_score(row)
        if score > best_score + 0.35 or (abs(score - best_score) <= 0.35 and (best_index is None or index < best_index)):
            best_index, best_score = index, score
    return best_index

def load_master_data(directory: str | None) -> dict[str, list[str]]:
    values: dict[str, set[str]] = {}
    if not directory or not Path(directory).is_dir():
        return {}
    for file in Path(directory).glob("*"):
        if file.suffix.lower() not in {".csv", ".xlsx", ".xlsm", ".xls"}:
            continue
        try:
            if file.suffix.lower() == ".csv":
                with file.open("r", encoding="utf-8-sig", newline="") as stream:
                    for row in csv.DictReader(stream):
                        for key, value in row.items():
                            text = clean_text(value)
                            if text:
                                values.setdefault(canonical_header(key or "value"), set()).add(text)
            else:
                for headers, data_rows in iter_tabular_file(file):
                    for row in data_rows:
                        for index, value in enumerate(row):
                            text = clean_text(value)
                            if text:
                                field = headers[index] if index < len(headers) else "value"
                                values.setdefault(field, set()).add(text)
        except Exception:
            # Optional master lists must not abort a batch; skip the bad file.
            continue
    return {key: sorted(items) for key, items in values.items()}


def iter_tabular_file(path: Path) -> Iterable[tuple[list[str], Iterable[Sequence[Any]]]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook
        book = load_workbook(path, read_only=True, data_only=True)
        try:
            for sheet in book.worksheets:
                rows = list(sheet.iter_rows(values_only=True))
                if not rows:
                    continue
                header_at = find_header_row(rows)
                headers = [canonical_header(str(value or f"column_{index + 1}")) for index, value in enumerate(rows[header_at])]
                yield headers, rows[header_at + 1 :]
        finally:
            book.close()
        return
    if suffix == ".xls":
        import xlrd
        book = xlrd.open_workbook(str(path))
        for sheet in book.sheets():
            rows = [sheet.row_values(index) for index in range(sheet.nrows)]
            if not rows:
                continue
            header_at = find_header_row(rows)
            headers = [canonical_header(str(value or f"column_{index + 1}")) for index, value in enumerate(rows[header_at])]
            yield headers, rows[header_at + 1 :]
        return
    with path.open("r", encoding="utf-8-sig", newline="") as stream:
        reader = csv.reader(stream)
        rows = list(reader)
        if not rows:
            return
        header_at = find_header_row(rows)
        headers = [canonical_header(str(value or f"column_{index + 1}")) for index, value in enumerate(rows[header_at])]
        yield headers, rows[header_at + 1 :]


@dataclass
class ValidationResult:
    original: str
    value: str
    changed: bool
    reason: str = ""
    confidence: float = 100.0
    ambiguous: bool = False
    review: bool = False


def validate_detailed(value: str, field: str, master: dict[str, list[str]]) -> ValidationResult:
    original = value
    reason = ""
    if DATE_FIELD.search(field):
        normalized = normalize_date(value)
        if normalized != value:
            value, reason = normalized, "date"
    elif NUMBER_FIELD.search(field) and not ID_FIELD.search(field):
        normalized = normalize_number(value)
        if normalized != value and re.search(r"\d", normalized):
            value, reason = normalized, "number"
    candidates = master.get(canonical_header(field), master.get("value", []))
    if not value or not candidates or value in candidates:
        return ValidationResult(original, value, value != original, reason, 100.0, False, value != original)
    close = get_close_matches(value, candidates, n=2, cutoff=0.88)
    if not close:
        return ValidationResult(original, value, value != original, reason, 100.0, False, value != original)
    best = SequenceMatcher(None, value.casefold(), close[0].casefold()).ratio()
    if len(close) > 1:
        second = SequenceMatcher(None, value.casefold(), close[1].casefold()).ratio()
        if best - second < 0.05:
            return ValidationResult(
                original, value, False, f"ambiguous:{close[0]}|{close[1]}", round(best * 100, 1), True, True
            )
    return ValidationResult(original, close[0], True, "master-data", round(best * 100, 1), False, True)


def validate_value(value: str, field: str, master: dict[str, list[str]]) -> tuple[str, bool]:
    result = validate_detailed(value, field, master)
    return result.value, result.review
