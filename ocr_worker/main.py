"""Offline batch OCR and structured-data cleaner used by the Vertex desktop app."""
from __future__ import annotations

import argparse
import json
import multiprocessing as mp
import os
import shutil
import sys
import tempfile
import time
import traceback
from concurrent.futures import ProcessPoolExecutor, as_completed
from pathlib import Path
from typing import Any

_WORKER_DIR = Path(__file__).resolve().parent
if str(_WORKER_DIR) not in sys.path:
    sys.path.insert(0, str(_WORKER_DIR))

from classify import classify_document
from clean import find_reliable_header_row, load_master_data, validate_value
from common import CONFIDENCE_THRESHOLD, FileResult, IMAGE_TYPES, SUPPORTED, TABULAR_TYPES, emit
from export import output_file, write_generic_tables, write_invoice, write_review_queue, write_status_report
from ingest import ingest
from invoice import invoice_table_is_reliable, parse_invoice_table
from ocr import image_pages, ocr_page_table
from tabular import clean_tabular
from templates import app_templates_path, load_templates, match_template, save_templates


def _process_invoice_ai(source: Path, master: dict[str, list[str]], output_dir: Path) -> FileResult:
    from invoice_ai import analyze as analyze_invoice_ai

    return analyze_invoice_ai(source, master, output_dir)


def app_log(level: str, message: str) -> None:
    directory = os.environ.get("VERTEX_LOG_DIR")
    if not directory:
        return
    try:
        Path(directory).mkdir(parents=True, exist_ok=True)
        with (Path(directory) / "worker.log").open("a", encoding="utf-8") as handle:
            handle.write(f"{time.strftime('%Y-%m-%dT%H:%M:%S')} {level} {message}\n")
    except OSError:
        pass


def expand_inputs(paths: list[str], output_dir: Path) -> list[Path]:
    files: list[Path] = []
    seen: set[str] = set()
    for raw in paths:
        item = Path(raw)
        if item.is_dir():
            candidates = [candidate for candidate in item.rglob("*") if candidate.is_file() and candidate.suffix.lower() in SUPPORTED]
        elif item.is_file():
            candidates = [item]
        else:
            continue
        for candidate in candidates:
            try:
                if output_dir in candidate.parents:
                    continue
            except ValueError:
                pass
            key = str(candidate.resolve()).casefold()
            if key in seen:
                continue
            seen.add(key)
            files.append(candidate)
    return files


def _clean_table_rows(rows: list[list[str]], scores: list[list[float]] | None, master: dict[str, list[str]]) -> dict:
    from clean import canonical_header, clean_text, find_reliable_header_row, validate_detailed
    if not rows:
        return {"name": "Extracted", "headers": [], "rows": []}
    width = max((len(row) for row in rows), default=0)
    first = [str(value or "").strip().casefold() for value in rows[0]] if rows else []
    # The local form analyser deliberately emits the stable two-column
    # ``Field | Value`` schema.  Do not let a generic header scorer promote a
    # later business field (for example ``Car Name | KIA``) into the header.
    header_at = 0 if first[:2] == ["field", "value"] else find_reliable_header_row(rows)
    use_as_header = header_at is not None
    if use_as_header:
        raw_headers = [str(value or "").strip() for value in list(rows[header_at]) + [""] * width][:width]
        data_rows = rows[header_at + 1 :]
        score_offset = header_at + 1
    else:
        raw_headers = [f"column_{index + 1}" for index in range(width)]
        data_rows = rows
        score_offset = 0
    raw_headers = [value or f"column_{index + 1}" for index, value in enumerate(raw_headers)]
    fields = [canonical_header(value) for value in raw_headers]
    cleaned_rows = []
    for offset, row in enumerate(data_rows):
        padded = list(row) + [""] * (width - len(row))
        padded = padded[:width]
        values, confidences, review, suggestion = [], [], False, ""
        review_columns: set[int] = set()
        if len(row) != width:
            review = True
            suggestion = f"expected {width} columns, row contains {len(row)} mapped cells"
        for index, value in enumerate(padded):
            field = fields[index] if index < len(fields) else "value"
            result = validate_detailed(clean_text(value, field), field, master)
            values.append(result.value)
            row_index = score_offset + offset
            conf = 100.0
            if scores and row_index < len(scores) and index < len(scores[row_index]):
                conf = float(scores[row_index][index] or 0)
            confidences.append(conf)
            if result.review:
                review = True
                suggestion = result.reason or suggestion
                review_columns.add(index)
            # Empty cells naturally have no OCR confidence.  They must not
            # turn an otherwise readable row into a manual-review row.
            if str(result.value).strip() and conf < CONFIDENCE_THRESHOLD:
                review = True
                review_columns.add(index)
        if any(str(value).strip() for value in values):
            if review and not review_columns:
                review_columns.add(next((index for index, value in enumerate(values) if str(value).strip()), 0))
            cleaned_rows.append({
                "values": values,
                "confidences": confidences,
                "review": review,
                "review_columns": sorted(review_columns),
                "suggestion": suggestion,
            })
    return {"name": "Extracted", "headers": raw_headers, "rows": cleaned_rows}


def _process_extracted(
    source: Path,
    tables: list[list[list[str]]],
    scores_list: list[list[list[float]]] | None,
    paragraphs: list[str],
    master: dict[str, list[str]],
    output_dir: Path,
    extra_warnings: list[str],
    force_generic: bool = False,
    local_form: bool = False,
) -> FileResult:
    destination = output_file(source, output_dir)
    blob = "\n".join(paragraphs) + "\n" + "\n".join(" ".join(cell for cell in row) for table in tables for row in table)
    column_count = max((len(row) for table in tables for row in table), default=0)
    row_count = sum(len(table) for table in tables)
    doc_type = classify_document(blob, column_count, row_count)
    known = load_templates(output_dir / "templates.json", app_templates_path())
    matched = match_template(blob, known)
    if matched and matched.get("type") in {"invoice", "table"}:
        doc_type = str(matched["type"])
    if local_form:
        doc_type = "form"
    records = low = 0
    review_items: list = []
    template = None
    warnings = list(extra_warnings)
    # Invoice fields must be earned by a recognisable item grid.  If OCR only
    # found page-wide text or merged cells, export the grid generically instead
    # of inventing a supplier, totals, or line items.
    merged = [row for table in tables for row in table]
    merged_scores = [row for scores in scores_list or [] for row in scores]
    if doc_type == "invoice" and not force_generic and invoice_table_is_reliable(merged):
        parsed = parse_invoice_table(merged, merged_scores or None, paragraphs)
        for item in parsed.get("items") or []:
            text, changed = validate_value(item.get("description", ""), "description", master)
            item["description"] = text
            if changed:
                item["review"] = True
                parsed["low_confidence"] = int(parsed.get("low_confidence") or 0) + 1
        records, low, review_items, template, destination = write_invoice(destination, source, parsed)
        if low:
            warnings.append("توجد عناصر صفراء تتطلب مراجعة.")
    else:
        if doc_type == "invoice":
            warnings.append("Invoice fields were not confidently detected; exported as a table for review.")
        generic = []
        for index, table in enumerate(tables or [[[line] for line in paragraphs]]):
            scores = scores_list[index] if scores_list and index < len(scores_list) else None
            payload = _clean_table_rows(table, scores, master)
            payload["name"] = f"Table{index + 1}" if len(tables) > 1 else "Extracted"
            generic.append(payload)
        # A fallback invoice must retain the page text that was not part of
        # the detected grid (supplier, dates, totals, and notes).  Keeping it
        # in a separate sheet is safer than mapping uncertain text to fields.
        if doc_type == "invoice":
            context_lines = [
                line.strip()
                for page in paragraphs
                for line in str(page).splitlines()
                if line.strip()
            ]
            if context_lines:
                generic.append({
                    "name": "OCR Context",
                    "headers": ["OCR text"],
                    "rows": [
                        {"values": [line], "confidences": [100.0], "review": False}
                        for line in context_lines
                    ],
                })
        records, low, review_items, template, destination = write_generic_tables(destination, source, generic)
        if low:
            warnings.append("توجد عناصر صفراء تتطلب مراجعة.")
    return FileResult(
        str(source), str(destination), records=records, low_confidence=low,
        warnings=warnings or None, template=template, review_items=review_items,
    )


def process_one(
    source_name: str,
    master: dict[str, list[str]],
    output_dir_name: str,
    mode: str = "clean",
) -> FileResult:
    source, output_dir = Path(source_name), Path(output_dir_name)
    suffix = source.suffix.lower()
    if suffix not in SUPPORTED:
        return FileResult(str(source), status="unsupported", error=f"صيغة غير مدعومة: {suffix or 'بدون امتداد'}")
    try:
        if str(mode or "clean").casefold() == "invoice_ai":
            if suffix in TABULAR_TYPES:
                # Spreadsheet invoices still use the deterministic tabular cleaner.
                return clean_tabular(source, master, output_dir)
            if suffix in IMAGE_TYPES or suffix == ".pdf":
                return _process_invoice_ai(source, master, output_dir)
            # Word/PowerPoint/text already carry their own text and layout. The
            # image pipeline cannot rasterise them, so they fall through to the
            # normal path rather than failing.
        if suffix in TABULAR_TYPES:
            return clean_tabular(source, master, output_dir)
        payload = ingest(source)
        if payload.get("kind") == "unsupported":
            return FileResult(str(source), status="unsupported", error="صيغة غير مدعومة.")
        tables = list(payload.get("tables") or [])
        paragraphs = list(payload.get("paragraphs") or [])
        images = list(payload.get("images") or [])
        warnings = list(payload.get("warnings") or [])
        scores_list: list[list[list[float]]] = []
        page_modes: list[str] = []
        image_iter = images
        if payload.get("stream_images"):
            image_iter = image_pages(source)
        for image in image_iter:
            table, scores, mode, page_text = ocr_page_table(image, include_page_text=True)
            if table:
                tables.append(table)
                scores_list.append(scores)
                page_modes.append(mode)
            if page_text.strip():
                paragraphs.append(page_text)
            del image
        if not tables and paragraphs:
            tables = [[[line] for line in paragraphs]]
        if not tables:
            return FileResult(str(source), status="failed", error="لم يتم استخراج بيانات من الملف.")
        force_generic = any(mode.startswith("local-ai-generic-table:") or mode.startswith("local-ai-form:") for mode in page_modes)
        if any(mode.startswith("local-ai-generic-table:") for mode in page_modes):
            warnings.append("تم تصدير الجدول للمراجعة؛ القيم غير المؤكدة لا تُعتمد كفاتورة نهائية.")
        local_form = bool(page_modes) and all(mode.startswith("local-ai-form:") for mode in page_modes)
        return _process_extracted(
            source,
            tables,
            scores_list or None,
            paragraphs,
            master,
            output_dir,
            warnings,
            force_generic=force_generic,
            local_form=local_form,
        )
    except Exception as error:
        app_log("ERROR", f"{source.name}: {type(error).__name__}")
        return FileResult(str(source), status="failed", error=f"{type(error).__name__}: {error}")


def cleanup_stale_artifacts(output_dir: Path | None = None) -> None:
    """Remove leftover sqlite sidecars and abandoned Office conversion folders."""
    temp = Path(tempfile.gettempdir())
    for path in temp.glob("vertex-office-*"):
        try:
            if path.is_dir():
                shutil.rmtree(path, ignore_errors=True)
        except OSError:
            pass
    for path in temp.glob("vertex-review-*.json"):
        try:
            path.unlink(missing_ok=True)
        except OSError:
            pass
    if output_dir and output_dir.is_dir():
        for path in output_dir.glob(".*.sqlite"):
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass


def default_output(inputs: list[str]) -> Path:
    first = Path(inputs[0])
    return (first if first.is_dir() else first.parent) / "VertexProcessed"


def apply_review_file(path: Path) -> None:
    from openpyxl import load_workbook
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    grouped: dict[str, list] = {}

    def pick(item: dict, *names: str) -> Any:
        lowered = {str(key).casefold(): value for key, value in item.items()}
        for name in names:
            if name.casefold() in lowered:
                return lowered[name.casefold()]
        return None

    for item in payload.get("items") or []:
        output = pick(item, "output")
        if not output:
            continue
        grouped.setdefault(str(output), []).append(item)
    for workbook_path, items in grouped.items():
        book = load_workbook(workbook_path)
        try:
            for item in items:
                sheet_name = pick(item, "sheet")
                row = pick(item, "row")
                column = pick(item, "column")
                if sheet_name is None or row is None or column is None:
                    continue
                book[str(sheet_name)].cell(int(row), int(column)).value = pick(item, "value")
            book.save(workbook_path)
        finally:
            book.close()


def run_self_check() -> int:
    checks: list[dict[str, Any]] = []

    def add(name: str, ok: bool, detail: str = "") -> None:
        checks.append({"name": name, "ok": ok, "detail": detail})

    modules = (
        ("cv2", "cv2"),
        ("PIL", "PIL"),
        ("numpy", "numpy"),
        ("openpyxl", "openpyxl"),
        ("xlrd", "xlrd"),
        ("pytesseract", "pytesseract"),
        ("pypdfium2", "pypdfium2"),
        ("docx", "docx"),
        ("pptx", "pptx"),
        ("striprtf", "striprtf"),
        ("paddle", "paddle"),
        ("paddleocr", "paddleocr"),
    )
    missing: list[str] = []
    for label, module_name in modules:
        try:
            __import__(module_name)
        except Exception as error:
            missing.append(f"{label}: {error}")
    add("Required Libraries", not missing, "PASS" if not missing else "; ".join(missing))

    tess = Path(os.environ.get("TESSERACT_CMD") or "")
    tessdata = Path(os.environ.get("TESSDATA_PREFIX") or "")
    tess_ok = tess.is_file()
    ara_ok = (tessdata / "ara.traineddata").is_file()
    eng_ok = (tessdata / "eng.traineddata").is_file()
    tess_detail = str(tess) if tess_ok else "tesseract.exe غير موجود"
    if tess_ok:
        import subprocess
        try:
            completed = subprocess.run(
                [str(tess), "--list-langs"],
                capture_output=True, text=True, timeout=20, env=os.environ.copy(),
            )
            tess_ok = completed.returncode == 0
            tess_detail = "tesseract --list-langs"
            langs = (completed.stdout or completed.stderr or "").casefold()
            ara_ok = ara_ok and "ara" in langs
            eng_ok = eng_ok and "eng" in langs
        except Exception as error:
            tess_ok = False
            tess_detail = str(error)
        try:
            from PIL import Image
            import pytesseract
            pytesseract.pytesseract.tesseract_cmd = str(tess)
            sample = Image.new("RGB", (80, 32), "white")
            pytesseract.image_to_string(sample, lang="eng")
            pytesseract.image_to_string(sample, lang="ara")
        except Exception as error:
            tess_ok = False
            tess_detail = str(error)
            ara_ok = False
            eng_ok = False
    add("Tesseract", tess_ok, tess_detail)
    add("Arabic OCR", ara_ok, str(tessdata / "ara.traineddata"))
    add("English OCR", eng_ok, str(tessdata / "eng.traineddata"))

    # PaddleOCR is the primary reader for images and scanned pages; Tesseract is
    # only the fallback. Build the engines for real so a broken model directory
    # is caught here rather than on the customer's first document.
    try:
        import perceive

        paddle_ok, paddle_detail = perceive.paddle_available()
        missing_models = [
            name for name, parts in (
                ("detection", ("det",)),
                ("Arabic recognition", ("rec", "arabic_PP-OCRv3_rec_infer")),
                ("English recognition", ("rec", "en_PP-OCRv3_rec_infer")),
            )
            if not perceive._model_dir(*parts)
        ]
        if missing_models:
            paddle_ok = False
            paddle_detail = "نماذج ناقصة: " + "، ".join(missing_models)
    except Exception as error:
        paddle_ok, paddle_detail = False, f"{type(error).__name__}: {error}"
    add("محرك القراءة (PaddleOCR)", paddle_ok, paddle_detail)

    ready = all(item["ok"] for item in checks)

    # Optional extras: report their state without failing the check.
    try:
        from invoice_understand import llm_available

        llm_ok, llm_detail = llm_available()
    except Exception as error:
        llm_ok, llm_detail = False, str(error)
    checks.append({
        "name": "نموذج التسمية (اختياري)",
        "ok": True,
        "detail": llm_detail if llm_ok else f"غير مثبّت — الاستخراج يعمل بدونه ({llm_detail})",
    })
    try:
        import vision

        vision_ok, vision_detail = vision.available()
    except Exception as error:
        vision_ok, vision_detail = False, str(error)
    checks.append({
        "name": "نموذج الرؤية (اختياري)",
        "ok": True,
        "detail": vision_detail if vision_ok else f"غير مثبّت — يُستخدم فقط للخلايا المشكوك فيها ({vision_detail})",
    })
    try:
        import paddle_vl

        ai_ok, ai_detail = paddle_vl.available()
    except Exception as error:
        ai_ok, ai_detail = False, f"{type(error).__name__}: {error}"
    checks.append({
        "name": "الاستخراج بالذكاء الاصطناعي (PaddleOCR-VL)",
        "ok": True,
        "detail": ai_detail if ai_ok else f"غير جاهز — يعمل المسار الهندسي بدلاً منه ({ai_detail})",
    })

    print(json.dumps({"ready": ready, "checks": checks}, ensure_ascii=False), flush=True)
    return 0 if ready else 2


def fetch_vl_models() -> int:
    """First-run download of the PaddleOCR-VL weights.

    Delegated to the VL virtual environment, because PaddleX owns the model
    registry and knows which archives the pinned version needs. Progress lines
    are forwarded to the desktop app unchanged.
    """
    import subprocess

    import paddle_vl

    interpreter = paddle_vl.worker_python()
    if interpreter is None:
        emit("مكوّنات PaddleOCR-VL غير مضمّنة في هذه النسخة.")
        return 2
    worker = _WORKER_DIR / "vl_worker.py"
    completed = subprocess.run(
        [str(interpreter), str(worker), "--fetch-models"],
        env=paddle_vl.worker_environment(),
    )
    if completed.returncode != 0:
        app_log("ERROR", "vl model fetch failed")
    return completed.returncode


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--request")
    parser.add_argument("--apply-review")
    parser.add_argument("--self-check", action="store_true")
    parser.add_argument("--fetch-vl-models", action="store_true")
    args = parser.parse_args()
    if args.self_check:
        return run_self_check()
    if args.fetch_vl_models:
        return fetch_vl_models()
    if args.apply_review:
        apply_review_file(Path(args.apply_review))
        return 0
    if not args.request:
        raise RuntimeError("مرّر --request أو --apply-review أو --self-check.")
    request = json.loads(Path(args.request).read_text(encoding="utf-8-sig"))

    def req(*names: str, default: Any = None) -> Any:
        lowered = {str(key).casefold(): value for key, value in request.items()}
        for name in names:
            if name.casefold() in lowered:
                return lowered[name.casefold()]
        return default

    input_paths = req("inputPaths", "InputPaths")
    if not input_paths:
        raise RuntimeError("طلب المعالجة لا يحتوي مسارات إدخال.")
    result_path = req("resultPath", "ResultPath")
    if not result_path:
        raise RuntimeError("طلب المعالجة لا يحتوي مسار النتيجة.")
    started = time.perf_counter()
    app_log("INFO", "processing start")
    output_dir = Path(req("outputDirectory", "OutputDirectory") or default_output(list(input_paths)))
    output_dir.mkdir(parents=True, exist_ok=True)
    cleanup_stale_artifacts(output_dir)
    files = expand_inputs(list(input_paths), output_dir)
    if not files:
        raise RuntimeError("لم يتم العثور على ملفات ضمن الاختيار.")
    master = load_master_data(req("masterDataDirectory", "MasterDataDirectory"))
    mode = str(req("mode", "Mode", default="clean") or "clean").casefold()
    if mode not in {"clean", "invoice_ai"}:
        mode = "clean"
    cpu = max(1, os.cpu_count() or 2)
    # Tesseract and OpenCV are memory-heavy.  Two concurrent files keeps the
    # self-contained build usable on ordinary customer laptops; callers can
    # still explicitly opt into a larger worker count.
    # Invoice AI (PaddleOCR) stays single-worker on typical 8GB machines.
    default_workers = 1 if mode == "invoice_ai" else min(2, cpu)
    requested = req("maxWorkers", "MaxWorkers") or default_workers
    workers = min(max(1, int(requested)), cpu, max(1, len(files)))
    if mode == "invoice_ai":
        workers = 1
        emit(f"تم العثور على {len(files):,} ملف. جاري فهم الفاتورة بالمحرك المحلي (عامل واحد)…")
    else:
        emit(f"تم العثور على {len(files):,} ملف. المعالجة المحلية المتوازية تعمل على {workers} نواة.")
    results: list[FileResult] = []
    if workers == 1:
        # Avoid ProcessPool spawn overhead/memory doubling on 8GB machines.
        for number, file in enumerate(files, start=1):
            try:
                result = process_one(str(file), master, str(output_dir), mode)
            except Exception as error:
                app_log("ERROR", f"worker {Path(file).name}: {type(error).__name__}")
                result = FileResult(str(file), status="failed", error=f"{type(error).__name__}: {error}")
            results.append(result)
            label = {"success": "تمت المعالجة", "failed": "فشل", "unsupported": "غير مدعوم"}.get(result.status, result.status)
            emit(f"({number}/{len(files)}) {Path(result.source).name}: {label}")
    else:
        with ProcessPoolExecutor(max_workers=workers, mp_context=mp.get_context("spawn")) as executor:
            futures = {
                executor.submit(process_one, str(file), master, str(output_dir), mode): file
                for file in files
            }
            for number, future in enumerate(as_completed(futures), start=1):
                source = futures[future]
                try:
                    result = future.result()
                except Exception as error:
                    app_log("ERROR", f"worker {Path(source).name}: {type(error).__name__}")
                    result = FileResult(str(source), status="failed", error=f"{type(error).__name__}: {error}")
                results.append(result)
                label = {"success": "تمت المعالجة", "failed": "فشل", "unsupported": "غير مدعوم"}.get(result.status, result.status)
                emit(f"({number}/{len(files)}) {Path(result.source).name}: {label}")
    if mode == "invoice_ai":
        # The VL model host is deliberately kept alive across the whole batch;
        # nothing after this point reads a page, so release its memory now.
        try:
            import paddle_vl

            paddle_vl.shutdown()
        except Exception as error:
            app_log("ERROR", f"vl shutdown: {type(error).__name__}")
    status = write_status_report(output_dir, results)
    review_path = write_review_queue(output_dir, results)
    save_templates(output_dir, [result.template for result in results if result.template])
    first_error = next((r.error for r in results if r.error), None)
    payload = {
        "outputPath": str(output_dir),
        "outputDirectory": str(output_dir),
        "statusReportPath": str(status),
        "reviewQueuePath": str(review_path),
        "elapsedSeconds": round(time.perf_counter() - started, 3),
        "filesProcessed": len(results),
        # A warning means the file was converted and needs review, not that it
        # failed to produce an Excel output.
        "filesSucceeded": sum(r.status == "success" for r in results),
        "filesWithWarnings": sum(r.status_label in {"auto_corrected", "needs_review"} for r in results),
        "filesFailed": sum(r.status in {"failed", "unsupported"} for r in results),
        "recordsExtracted": sum(r.records for r in results),
        "lowConfidenceItems": sum(r.low_confidence for r in results),
        "firstError": first_error,
    }
    Path(str(result_path)).write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
    cleanup_stale_artifacts(output_dir)
    app_log("INFO", f"processing end records={payload['recordsExtracted']} failed={payload['filesFailed']}")
    emit(f"تمت معالجة ونقل {payload['recordsExtracted']:,} بنداً بدقة خلال {payload['elapsedSeconds']} ثانية.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as error:
        traceback.print_exception(error, file=sys.stderr)
        raise SystemExit(1)
