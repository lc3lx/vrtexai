"""The page, reproduced in Excel — a scan turned into a spreadsheet.

The customer's test is simple and it is the right one: put the invoice next to
the workbook and they should read the same, top to bottom, in the same order,
with every piece of data in its own cell. Nothing on the paper missing, nothing
invented, and tidier than the paper.

That is a different job from extraction, and the two used to be confused. The
workbook was built out of what the reader *understood* — a line-item table with
the header fields repeated beside it — so everything the reader had no name for
simply never appeared: the company, the tax number, the invoice title, the
summary box, the amount in words, the notes, the signatures. Two thirds of the
document was missing and the third that survived had been rearranged.

So this module writes the page, block by block, in printed order. The
interpretation still happens — it is what puts real numbers and live formulas in
the item table, and what flags a figure that does not add up — but it decorates
the page rather than replacing it. :func:`excel_builder.write_ai_workbook` adds
the flat data sheet after this one, for the customer who wants to pivot rather
than read.
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Sequence

# One label and its value, inside a run of text. The page separates fields
# printed side by side with a run of white space, and that is the only signal
# that survives into a line of characters.
_CHUNKS = re.compile(r"\s{2,}|\t|\|")
_LABEL = re.compile(r"^([^:：]{1,40}?)\s*[:：]\s*(.*)$")
_CLOCK = re.compile(r"\d\s*[:：]\s*\d")

MAX_TEXT_CELLS = 8


def split_line(line: str) -> list[tuple[str, bool]]:
    """One printed line as ``(text, is_label)`` cells, left to right.

    "Invoice No: 1    Date: 12/27/2021" is four pieces of information, and the
    customer asked for each in its own cell. A colon between two digits is a
    clock, not a label, so "Time 3:00 PM" stays one cell.
    """
    cells: list[tuple[str, bool]] = []
    for chunk in _CHUNKS.split(str(line or "")):
        chunk = chunk.strip()
        if not chunk:
            continue
        match = None if _CLOCK.search(chunk) else _LABEL.match(chunk)
        if match and match.group(1).strip():
            cells.append((match.group(1).strip(), True))
            value = match.group(2).strip()
            if value:
                cells.append((value, False))
        else:
            cells.append((chunk, False))
    return cells[:MAX_TEXT_CELLS]


def _section_width(section: dict[str, Any], document: dict[str, Any]) -> int:
    kind = section.get("kind")
    if kind == "items":
        from excel_builder import plan_columns

        return max(len(plan_columns(document)), 1)
    if kind == "table":
        rows = section.get("rows") or []
        return max([len(section.get("columns") or [])] + [len(row) for row in rows] + [1])
    if kind == "text":
        return max((len(split_line(line)) for line in section.get("lines") or []), default=1)
    if kind in {"totals", "fields"}:
        return 2
    return 1


def sheet_width(document: dict[str, Any]) -> int:
    """How many columns the page needs, measured across every block on it."""
    sections = document.get("sections") or []
    return max([_section_width(section, document) for section in sections] + [2])


def write_document(
    sheet,
    document: dict[str, Any],
    source: Path,
    destination: Path,
    styles: dict[str, Any],
) -> tuple[int, int, list[dict[str, Any]]]:
    """Write one page into one sheet. Returns (records, flagged, review items)."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, Side

    import excel_builder as builder

    thin = styles["thin"]
    yellow = styles["yellow"]
    band = styles["band"]
    label_fill = styles["label_fill"]
    stripe = styles.get("stripe")

    rtl = str(document.get("direction") or "ltr").casefold() == "rtl"
    sheet.sheet_view.rightToLeft = rtl
    # No gridlines: the page's own rules are drawn where the page has them, and
    # a grid behind them makes the sheet look like a spreadsheet rather than a
    # copy of the document.
    sheet.sheet_view.showGridLines = False
    say = builder.document_words(document)

    currency = builder.money_format(str(document.get("currency") or ""))
    quantity_format = builder.FRACTION_FORMAT
    width = sheet_width(document)
    widths: dict[int, int] = {}
    review_items: list[dict[str, Any]] = []
    records = flagged = 0
    row = 1

    def track(column: int, text: Any) -> None:
        builder._track(widths, column, text)

    def flag(at_row: int, column: int, heading: str, value: Any, note: str) -> None:
        review_items.append({
            "output": str(destination),
            "sheet": sheet.title,
            "row": at_row,
            "column": column,
            "header": heading,
            "value": "" if value is None else value,
            "confidence": "",
            "suggestion": note,
        })

    def merge_to_width(at_row: int, start: int) -> None:
        if width > start:
            sheet.merge_cells(start_row=at_row, start_column=start, end_row=at_row, end_column=width)

    # ---- the blocks of the page -----------------------------------------
    def write_title(text: str) -> None:
        nonlocal row
        cell = sheet.cell(row, 1, text)
        cell.font = Font(name=builder.FONT_NAME, size=14, bold=True, color=builder.TITLE_TEXT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        merge_to_width(row, 1)
        sheet.row_dimensions[row].height = 24
        track(1, text[:40])
        row += 1

    def write_text(lines: Sequence[str]) -> None:
        nonlocal row
        for line in lines:
            cells = split_line(line)
            if not cells:
                continue
            column = 1
            for text, is_label in cells:
                cell = sheet.cell(row, column, text)
                builder._style_cell(cell, border=thin, bold=is_label)
                if is_label:
                    cell.fill = label_fill
                cell.alignment = builder._text_alignment(
                    text, wrap=len(text) > builder.WRAP_AT
                )
                track(column, text)
                column += 1
            # A lone value, or a single label and its value, is stretched to the
            # edge of the sheet so the block reads as one line rather than as a
            # column of stubs.
            if column - 1 <= 2:
                merge_to_width(row, column - 1)
            row += 1

    def write_table(columns: Sequence[str], rows: Sequence[Sequence[str]]) -> None:
        nonlocal row
        if columns:
            for index, heading in enumerate(columns, start=1):
                cell = sheet.cell(row, index, heading)
                cell.font = Font(name=builder.FONT_NAME, size=11, bold=True,
                                 color=builder.BAND_TEXT)
                cell.fill = band
                cell.border = thin
                cell.alignment = Alignment(horizontal="center", vertical="center",
                                           wrap_text=True)
                track(index, heading)
            sheet.row_dimensions[row].height = 20
            row += 1
        start = row
        for values in rows:
            for index, text in enumerate(values, start=1):
                number = builder._as_number(text) if str(text).strip() else None
                strict = _plain_number(text)
                cell = sheet.cell(row, index, strict if strict is not None else text)
                builder._style_cell(cell, border=thin)
                if stripe is not None and (row - start) % 2 == 1:
                    cell.fill = stripe
                if strict is not None:
                    cell.number_format = quantity_format
                    cell.alignment = builder._number_alignment()
                    track(index, builder._shown(strict))
                else:
                    cell.alignment = builder._text_alignment(
                        text, wrap=len(str(text)) > builder.WRAP_AT
                    )
                    track(index, text)
                del number
            row += 1

    def write_items() -> None:
        nonlocal row, records, flagged
        fields = builder.plan_columns(document)
        items = list(document.get("items") or [])
        if not fields or not items:
            return
        numeric_fields = builder.numeric_text_fields(items, fields)
        percent_fields = {
            field for field, heading, role in fields
            if builder.percent_column(heading, field, role, [i.get(field) for i in items])
        }
        # A column of whole numbers is shown without a decimal point: Excel
        # renders 10 as "10." under a format that allows decimals.
        counts = {
            field: builder.quantity_format_for([item.get(field) for item in items])
            for field, _heading, _role in fields
        }

        header_row = row
        role_columns: dict[str, int] = {}
        for index, (_field, heading, role) in enumerate(fields, start=1):
            cell = sheet.cell(row, index, heading)
            cell.font = Font(name=builder.FONT_NAME, size=11, bold=True, color=builder.BAND_TEXT)
            cell.fill = band
            cell.border = thin
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            track(index, heading)
            if role != "other" and role not in role_columns:
                role_columns[role] = index
        sheet.row_dimensions[row].height = 20
        row += 1

        first = row
        qty_column = role_columns.get("qty")
        price_column = role_columns.get("unit_price")
        total_column = role_columns.get("line_total")
        can_compute = bool(qty_column and price_column and total_column)

        for item in items:
            reviews = item.get("review") or {}
            notes = item.get("notes") or {}
            painted = False
            for index, (field, heading, role) in enumerate(fields, start=1):
                value = item.get(field)
                numeric = role in builder.MONEY_ROLES or role in builder.QTY_ROLES
                if field in numeric_fields:
                    converted = builder._as_number(value)
                    if converted is not None:
                        value, numeric = converted, True

                is_formula = False
                if can_compute and role == "line_total":
                    qty, price = item.get("qty"), item.get("unit_price")
                    if qty is not None and price is not None:
                        letter = builder._column_letter
                        cell = sheet.cell(
                            row, index, f"={letter(qty_column)}{row}*{letter(price_column)}{row}"
                        )
                        is_formula = True
                        if value is not None:
                            cell.comment = Comment(f"{say('read_value')}: {value:,.2f}", "Vertex")
                        track(index, builder._shown(value if value is not None else qty * price))
                    else:
                        cell = sheet.cell(row, index, value)
                        track(index, builder._shown(value))
                elif numeric:
                    cell = sheet.cell(row, index, value)
                    track(index, builder._shown(value))
                else:
                    text = str(value).strip() if value not in (None, "") else ""
                    cell = sheet.cell(row, index, text)
                    track(index, text)

                builder._style_cell(cell, border=thin)
                if stripe is not None and (row - first) % 2 == 1:
                    cell.fill = stripe
                if field in percent_fields:
                    cell.number_format = builder.PERCENT_FORMAT
                    cell.alignment = builder._number_alignment()
                elif role in builder.MONEY_ROLES:
                    cell.number_format = currency
                    cell.alignment = builder._number_alignment()
                elif role in builder.QTY_ROLES:
                    cell.number_format = counts.get(field, quantity_format)
                    cell.alignment = builder._number_alignment(horizontal="center")
                elif numeric:
                    cell.number_format = counts.get(field, quantity_format)
                    cell.alignment = builder._number_alignment()
                else:
                    cell.alignment = builder._text_alignment(
                        cell.value, wrap=len(str(cell.value or "")) > builder.WRAP_AT
                    )

                note = notes.get(field) or notes.get(role)
                if note and not is_formula:
                    cell.comment = Comment(str(note), "Vertex")
                if reviews.get(field) or reviews.get(role):
                    cell.fill = yellow
                    painted = True
                    if not is_formula:
                        flag(row, index, heading, cell.value, str(note or ""))
                    elif qty_column:
                        flag(row, qty_column, say("qty"), item.get("qty"), str(note or ""))
            records += 1
            if painted:
                flagged += 1
            row += 1

        last = row - 1
        sheet.auto_filter.ref = f"A{header_row}:{builder._column_letter(len(fields))}{last}"

        # The line the page prints under its own table — "مجموع | 20 | 144,400"
        # — written where the page puts it, in bold, above the totals block.
        for footing in document.get("item_totals") or []:
            values = footing.get("values") or {}
            for index, (field, _heading, role) in enumerate(fields, start=1):
                text = str(values.get(field) or "").strip()
                number = _plain_number(text)
                cell = sheet.cell(row, index, number if number is not None else text)
                builder._style_cell(cell, border=thin, bold=True)
                cell.fill = label_fill
                if number is not None:
                    cell.number_format = (
                        builder.PERCENT_FORMAT if field in percent_fields
                        else currency if role in builder.MONEY_ROLES
                        else quantity_format
                    )
                    cell.alignment = builder._number_alignment()
                else:
                    cell.alignment = builder._text_alignment(text)
                track(index, text)
            row += 1

        state["items"] = (first, last, role_columns, len(fields))

    def write_totals() -> None:
        nonlocal row, flagged
        totals = document.get("totals") or {}
        review = document.get("totals_review") or {}
        notes = document.get("totals_notes") or {}
        first, last, role_columns, item_width = state.get("items", (0, 0, {}, 0))
        # A page can carry a column of amounts and print no sum of them — a
        # shipping manifest usually does. The sheet adds them up rather than
        # leaving the customer to, as a formula and under a label that says it
        # was calculated, so nobody mistakes it for a figure off the paper.
        if not totals and not (first and last and role_columns.get("line_total")):
            return
        total_column = role_columns.get("line_total") or max(2, min(width, item_width or width))
        label_column = max(1, total_column - 1)
        written: dict[str, int] = {}

        for key in builder._TOTAL_ORDER:
            if key not in totals and not (key == "subtotal" and first):
                continue
            value = totals.get(key)
            printed = key in totals
            label_cell = sheet.cell(
                row, label_column,
                say(key, key) if printed else say("computed_total"),
            )
            is_grand = key == "grand_total"
            builder._style_cell(label_cell, border=thin, bold=True)
            label_cell.fill = label_fill
            label_cell.alignment = Alignment(horizontal="left" if rtl else "right")
            track(label_column, label_cell.value)

            formula = None
            letter = builder._column_letter(total_column)
            if key == "subtotal" and first and last and role_columns.get("line_total"):
                formula = f"=SUM({letter}{first}:{letter}{last})"
            elif key == "tax_amount" and "subtotal" in written and totals.get("tax_rate"):
                formula = f"={letter}{written['subtotal']}*{float(totals['tax_rate'])}"
            elif is_grand and "subtotal" in written:
                parts = [f"{letter}{written['subtotal']}"]
                if "tax_amount" in written:
                    parts.append(f"+{letter}{written['tax_amount']}")
                if "discount" in written:
                    parts.append(f"-ABS({letter}{written['discount']})")
                formula = "=" + "".join(parts)

            value_cell = sheet.cell(row, total_column, formula if formula else value)
            builder._style_cell(value_cell, border=thin, bold=True, size=12 if is_grand else 11)
            value_cell.number_format = currency
            value_cell.alignment = builder._number_alignment()
            if formula is not None and value is not None:
                value_cell.comment = Comment(f"{say('read_value')}: {value:,.2f}", "Vertex")
            track(total_column, builder._shown(value))
            if is_grand:
                value_cell.border = Border(
                    left=thin.left, right=thin.right,
                    top=Side(style="thin", color=builder.BAND),
                    bottom=Side(style="double", color=builder.BAND),
                )
            note = notes.get(key)
            if note and formula is None:
                value_cell.comment = Comment(str(note), "Vertex")
            if review.get(key):
                value_cell.fill = yellow
                flagged += 1
                if formula is None:
                    flag(row, total_column, str(label_cell.value), value, str(note or ""))
            written[key] = row
            row += 1

    def write_fields(fields: dict[str, Any]) -> None:
        """The header fields, one per row, when the page itself was not kept."""
        nonlocal row
        for key, value in fields.items():
            label = say(str(key).casefold(), str(key))
            label_cell = sheet.cell(row, 1, label)
            builder._style_cell(label_cell, border=thin, bold=True)
            label_cell.fill = label_fill
            label_cell.alignment = builder._text_alignment(label)
            value_cell = sheet.cell(row, 2, value)
            builder._style_cell(value_cell, border=thin)
            value_cell.alignment = builder._text_alignment(
                value, wrap=len(str(value)) > builder.WRAP_AT
            )
            merge_to_width(row, 2)
            track(1, label)
            track(2, value)
            row += 1

    state: dict[str, Any] = {}
    sections = list(document.get("sections") or [])
    if not sections:
        # Nothing recorded what the page looked like — an older payload, or a
        # reader that returned only a table. Write what there is, in the order a
        # document has it.
        sections = [{"kind": "fields"}, {"kind": "items"}, {"kind": "totals"}]

    # Which block, if any, is the page's masthead. A transcribing model labels
    # every paragraph "text", so nothing arrives marked as the title; the first
    # short line on the page is it, which is where a document puts its name.
    masthead = -1
    for number, section in enumerate(sections):
        if section.get("kind") != "text":
            continue
        lines = [str(line) for line in section.get("lines") or []]
        if len(lines) == 1 and len(lines[0]) <= 60 and len(split_line(lines[0])) == 1:
            masthead = number
        break

    for number, section in enumerate(sections):
        kind = section.get("kind")
        before = row
        if kind == "title":
            write_title(str(section.get("text") or ""))
        elif kind == "text" and number == masthead:
            write_title(str((section.get("lines") or [""])[0]))
        elif kind == "text":
            write_text([str(line) for line in section.get("lines") or []])
        elif kind == "table":
            write_table(
                [str(name) for name in section.get("columns") or []],
                [[str(value) for value in values] for values in section.get("rows") or []],
            )
        elif kind == "fields":
            write_fields(document.get("header") or {})
        elif kind == "items":
            write_items()
        elif kind == "totals":
            write_totals()
        if row > before:
            row += 1  # one blank line between blocks, as on the page

    if row == 1:
        sheet.cell(1, 1, say("no_page_data"))
    for column, size in widths.items():
        sheet.column_dimensions[builder._column_letter(column)].width = max(
            builder.MIN_WIDTH, size
        )

    sheet.page_setup.orientation = "landscape" if width > 6 else "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    return records, flagged, review_items


def _plain_number(text: Any) -> float | None:
    """A cell that is only a number, so a table's figures stay numbers."""
    from table_shape import numeric_cell

    return numeric_cell(text)
