"""Professional Excel output for the AI-led path.

Three things separate this from :mod:`export`:

* **One flat table.** The document's header — shipper, consignee, invoice
  number, dates — is written as *columns beside the line items*, repeated on
  every row, not as a stack of label/value pairs above the table in columns A
  and B. A sheet shaped that way is a data table: it sorts, it filters, it
  pivots, and a row still says who shipped what after the sort. The stacked
  form looked like the paper but was useless to work with, which is what the
  customer's accounts office actually does with the file.
* **Real formulas.** A line total is written as ``=B5*C5`` and a subtotal as
  ``=SUM(D5:D12)``, so the sheet recalculates when the customer edits a
  quantity. The number the model actually read is not thrown away — it is
  attached to the cell as a comment, and the cell is highlighted when the two
  disagree. That way the formula is live and the evidence is still there.
* **Formatting.** Arial throughout, a coloured bold header band, thin borders
  on every table cell, numbers written as numbers with currency and quantity
  formats, text aligned to the side its own script starts from, and column
  widths measured from the content that is actually in them.

One sheet per document, one workbook per source file. Pages of the same order
are merged upstream (:func:`ai_extract.merge_pages`), so a five-page manifest is
one continuous table, not five sheets each repeating the same heading.
"""
from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Sequence

from common import YELLOW
from export import _save_workbook
from templates import build_template

FONT_NAME = "Arial"
BAND = "1F4E79"       # header band
BAND_TEXT = "FFFFFF"
RULE = "B4C6E7"       # table grid
STRIPE = "F2F7FC"     # banded rows, a tint of the header band
LABEL_FILL = "DDEBF7"  # field labels
TITLE_TEXT = "1F4E79"

# Roles the builder gives special treatment. Anything else is a text column.
MONEY_ROLES = {"unit_price", "line_total", "discount", "tax"}
QTY_ROLES = {"qty"}

# What an empty text cell says. A field the page left blank — a shipper phone
# nobody filled in — is written rather than skipped, so the row keeps its shape
# and the columns underneath stay lined up. Numbers stay genuinely empty: "N/A"
# in a money column is text, and text in the middle of a column is what stops
# SUM and AVERAGE from being trustworthy.
MISSING = "N/A"

# Ceilings, not targets. The header of a real invoice carries a dozen fields;
# past this it is a page of prose being pushed sideways, and a sheet two hundred
# columns wide helps nobody.
MAX_HEADER_COLUMNS = 20
MAX_WIDTH = 55
MIN_WIDTH = 10
WRAP_AT = 45

# Every word this builder puts in a cell, in both languages: (Arabic, English).
#
# The document chooses, not the product. An English invoice that comes back with
# "الإجمالي" over its figures has been translated by the tool rather than
# transcribed, and the sheet no longer matches the paper it came from — which is
# the one thing an accounts office checks it against. The same decision that
# sets the sheet's direction picks the column here, so a page is never
# right-to-left with English headings or the other way round.
_TEXT: dict[str, tuple[str, str]] = {
    # header fields
    "supplier": ("المورد", "Supplier"),
    "client_name": ("العميل", "Customer"),
    "shipper": ("المرسِل", "Shipper"),
    "shipper_phone": ("هاتف المرسِل", "Shipper phone"),
    "shipper_address": ("عنوان المرسِل", "Shipper address"),
    "consignee": ("المرسَل إليه", "Consignee"),
    "consignee_phone": ("هاتف المرسَل إليه", "Consignee phone"),
    "consignee_address": ("عنوان المرسَل إليه", "Consignee address"),
    "invoice_number": ("رقم الفاتورة", "Invoice no."),
    "purchase_order": ("أمر الشراء", "Purchase order"),
    "invoice_date": ("تاريخ الفاتورة", "Invoice date"),
    "due_date": ("تاريخ الاستحقاق", "Due date"),
    "tax_number": ("الرقم الضريبي", "Tax number"),
    "payment_terms": ("شروط الدفع", "Payment terms"),
    # item columns
    "description": ("الوصف", "Description"),
    "sku": ("رمز الصنف", "SKU"),
    "qty": ("الكمية", "Qty"),
    "unit_price": ("سعر الوحدة", "Unit price"),
    "line_total": ("الإجمالي", "Line total"),
    "discount": ("الخصم", "Discount"),
    "tax": ("الضريبة", "Tax"),
    "unit": ("الوحدة", "Unit"),
    "date": ("التاريخ", "Date"),
    # totals
    "subtotal": ("المجموع الفرعي", "Subtotal"),
    "tax_amount": ("الضريبة", "Tax"),
    "tax_rate": ("نسبة الضريبة", "Tax rate"),
    "grand_total": ("الإجمالي النهائي", "Total"),
    "amount_paid": ("المدفوع", "Paid"),
    # page furniture
    "page_n": ("صفحة", "Page"),
    "read_value": ("القيمة المقروءة من الصورة", "Value read from the image"),
    "no_page_data": (
        "لم يتم استخراج أي بيانات من هذه الصفحة.",
        "No data could be extracted from this page.",
    ),
    "no_file_data": (
        "لم يتم استخراج أي بيانات من هذا الملف.",
        "No data could be extracted from this file.",
    ),
    # sheet names
    "data_sheet": ("البيانات", "Data"),
    # review sheet
    "review": ("المراجعة", "Review"),
    "review_title": ("ملخّص المراجعة", "Review summary"),
    "source": ("المصدر", "Source"),
    "file": ("الملف", "File"),
    "pages": ("الصفحات", "Pages"),
    "items_extracted": ("البنود المستخرجة", "Items extracted"),
    "what_checked": ("ما جرى التحقق منه", "What was checked"),
    "shape": ("الشكل", "Shape"),
    "shape_note": (
        "الأرقام أرقام والأدوار من مجموعة معروفة",
        "Numbers are numbers, and every column role comes from a known set",
    ),
    "arithmetic": ("الحساب", "Arithmetic"),
    "arithmetic_note": (
        "الكمية × السعر = الإجمالي، ومجموع البنود = المجموع الفرعي",
        "Quantity x price = line total, and the item sum = the subtotal",
    ),
    "pixels": ("البكسلات", "Pixels"),
    "pixels_note": (
        "كل رقم له ما يقابله في قراءة مستقلة للصورة",
        "Every figure has a match in an independent reading of the image",
    ),
    # Said plainly when the check did not run. A page nothing verified must not
    # read like a page that was verified and found clean.
    "pixels_off": (
        "لم تجرِ قراءة مستقلة لهذه الصفحة — لم يُقابَل أي رقم بالصورة",
        "No independent reading was available — no figure was checked against the image",
    ),
    # Said differently from a printed total, because it is not one: the page
    # carried a column of amounts and no sum, so the sheet adds them up. Naming
    # it plainly is what keeps the workbook a record rather than a claim.
    "computed_total": ("المجموع (محسوب)", "Total (calculated)"),
    "repaired": ("قيم صُحّحت بالحساب", "Corrected by arithmetic"),
    "repaired_note": (
        "رقم قُرئ خطأً وأثبت الحساب قيمته الصحيحة؛ الخلية مُعلَّمة وفيها ملاحظة بما تغيّر",
        "A misread figure whose correct value the arithmetic proved; the cell is "
        "highlighted and carries a note saying what changed",
    ),
    "needs_review": ("يحتاج مراجعة", "Needs review"),
    "all_clean": (
        "لا شيء. كل قيمة لها دليل في الصورة وحسابها صحيح.",
        "Nothing. Every value has evidence in the image and its arithmetic checks out.",
    ),
    "col_sheet": ("الورقة", "Sheet"),
    "col_cell": ("الخلية", "Cell"),
    "col_value": ("القيمة", "Value"),
    "col_reason": ("السبب", "Reason"),
    "needs_confirming": ("تحتاج تأكيداً", "Needs confirming"),
}

# Written in the order a document states them. ``amount_paid`` sits after the
# total because that is where a receipt prints it — it is what was handed over,
# not part of what was owed, so no formula is built from it.
_TOTAL_ORDER = ("subtotal", "discount", "tax_amount", "grand_total", "amount_paid")

# The order the header columns are laid out in, when the document happens to
# carry them. Who and what first, then the references, then the dates and terms
# — the order somebody reads a shipment in. Fields this product has no name for
# keep their printed label and follow, in the order the page printed them.
_HEADER_ORDER = (
    "supplier", "shipper", "shipper_phone", "shipper_address",
    "client_name", "consignee", "consignee_phone", "consignee_address",
    "invoice_number", "purchase_order", "tax_number",
    "invoice_date", "due_date", "payment_terms",
)

# Keys the item objects carry for the builder's own use, never as a column.
_PRIVATE_ITEM_KEYS = {"review", "notes"}

# A name this product invented because the page gave the column none.
_AUTOMATIC_NAME = re.compile(r"^column_\d+$", re.I)


def words_for(direction: str):
    """A lookup that answers in the language the document is written in."""
    index = 0 if str(direction or "ltr").casefold() == "rtl" else 1

    def say(key: str, default: str | None = None) -> str:
        pair = _TEXT.get(key)
        if pair is None:
            return default if default is not None else key
        return pair[index]

    return say


def document_words(document: dict[str, Any]):
    return words_for(str(document.get("direction") or "ltr"))

_CURRENCY_SYMBOLS = {
    "SAR": "ر.س", "AED": "د.إ", "QAR": "ر.ق", "KWD": "د.ك", "EGP": "ج.م",
    "JOD": "د.أ", "USD": "$", "EUR": "€", "GBP": "£",
}


# Currencies whose symbol is written before the amount. $1,200.00 is right and
# 1,200.00 $ is not; ر.س goes after its number and $ does not, so the symbol's
# side is a property of the currency rather than a house style.
_SYMBOL_LEADS = {"USD", "EUR", "GBP", "CAD", "AUD", "CHF", "JPY", "CNY", "INR"}


def money_format(currency: str) -> str:
    """An Excel number format carrying the document's own currency."""
    code = (currency or "").strip().upper()
    symbol = _CURRENCY_SYMBOLS.get(code, (currency or "").strip())
    # Quotes and semicolons would end the literal and corrupt the format string.
    symbol = re.sub(r'["\\;\[\]]', "", symbol)[:6]
    if not symbol:
        return "#,##0.00"
    if code in _SYMBOL_LEADS:
        return f'"{symbol}"#,##0.00'
    return f'#,##0.00" {symbol}"'


def _column_letter(index: int) -> str:
    from openpyxl.utils import get_column_letter

    return get_column_letter(index)


def plan_columns(document: dict[str, Any]) -> list[tuple[str, str, str]]:
    """Decide the item table's columns as ``(field, heading, role)``.

    The item objects carry the data, but the model also reports the headings it
    saw printed and a role for each. Those headings are used when they can be
    matched to a role, so the sheet reads like the original document instead of
    like the schema.
    """
    from ai_extract import ROLES

    say = document_words(document)
    headings: dict[str, str] = {}
    order: list[str] = []
    roles = document.get("column_roles") or []
    columns = document.get("columns") or []
    for index, role in enumerate(roles):
        if role in ROLES and role != "other" and role not in headings:
            heading = str(columns[index]).strip() if index < len(columns) else ""
            # The heading actually printed on the page wins over our own word for
            # it. Falling back to a translation is the last resort, not the plan.
            headings[role] = heading or say(role, role)
            order.append(role)

    fields: list[tuple[str, str, str]] = []
    seen: set[str] = set()

    def add(field: str) -> None:
        key = field.casefold()
        if key in seen:
            return
        seen.add(key)
        role = key if key in ROLES else "other"
        heading = headings.get(role) or say(role, field)
        fields.append((field, heading, role))

    items = list(document.get("items") or [])
    present: list[str] = []
    for item in items:
        for key in item:
            if key in _PRIVATE_ITEM_KEYS or key.startswith("_") or key in present:
                continue
            present.append(key)

    # A column with no heading of its own and nothing in it on any row. Pages
    # are drawn with a margin column and a reader transcribes it as real, so the
    # table arrived one column wider than it is — every heading shifted right of
    # its data, under a name like "column_1" that came from nowhere the customer
    # could see. A column the document named is kept even when it is empty; one
    # only we named is not a column.
    def is_phantom(field: str) -> bool:
        if not _AUTOMATIC_NAME.match(field.strip()):
            return False
        return not any(str(item.get(field) or "").strip() for item in items)

    present = [field for field in present if not is_phantom(field)]
    known = {field.casefold(): field for field in present}

    # The order the document printed, column by column, taken from the reader's
    # own list rather than from the order the keys happen to sit in an object.
    # It is the customer's choice: the sheet is meant to read like the paper it
    # came from, so a recognised column is not promoted to the front for being
    # recognised — "Inward Quantity" stays where it was printed even though the
    # column beside it is the one the formulas use.
    for index, role in enumerate(roles):
        name = str(columns[index]).strip() if index < len(columns) else ""
        key = role if role in ROLES and role != "other" else name
        field = known.get(str(key).casefold())
        if field is not None:
            add(field)
    for field in present:
        add(field)
    return fields


def plan_header_columns(document: dict[str, Any]) -> list[tuple[str, str]]:
    """The header fields as ``(key, heading)`` columns beside the item detail.

    The customer's complaint was precise: shipper and consignee arriving as a
    vertical stack in columns A and B, above a table they had nothing to do
    with. Read like that the sheet is a picture of the page; laid out as columns
    it is a table, and the line "who sent this, to whom, on which order" travels
    with every row through a sort or a filter.
    """
    say = document_words(document)
    header = document.get("header") or {}
    keys = [str(key).strip() for key in header if str(key).strip()]
    known = [key for name in _HEADER_ORDER for key in keys if key.casefold() == name]
    rest = [key for key in keys if key not in known]

    # One column per value, not one per label. A page that prints its tax number
    # twice — once labelled "TRN" and once "الرقم الضريبي" — was giving the sheet
    # two identical columns repeated down every row, and a badly read page gave
    # it a dozen.
    chosen: list[str] = []
    seen: set[str] = set()
    for key in known + rest:
        value = re.sub(r"\W+", "", str(header.get(key) or "").casefold())
        if value and value in seen:
            continue
        seen.add(value)
        chosen.append(key)
    return [(key, say(key.casefold(), key)) for key in chosen[:MAX_HEADER_COLUMNS]]


# Headings whose digits identify something rather than measure it. A tracking
# number that Excel helpfully turns into 1.02442E+11, or a phone number that
# loses its leading zero, is a defect — so these columns stay text however
# numeric they look.
_IDENTIFIER_HEADING = re.compile(
    r"sku|code|no\.?\b|number|ref\b|serial|\bid\b|phone|mobile|tel\b|fax|zip|postal"
    r"|barcode|iban|account|awb|tracking"
    r"|رقم|هاتف|جوال|كود|مرجع|حساب|بوليصة",
    re.I,
)

# A rate, not an amount. The column is stored as the fraction it means and
# shown back the way the page printed it.
# Every format here spells its decimals out with zeros rather than with the "#"
# placeholder. "#,##0.###" is correct Excel and reads beautifully in Excel — and
# the phone viewer the customer opened the file on rendered it literally, so a
# column of net amounts came out as "7.###", "8.###", "211.###" in front of
# their own client. A format that is right everywhere beats one that is elegant
# in one program.
PERCENT_FORMAT = "0.00%"

# Excel prints a trailing point for a whole number under "#,##0.###" — a
# quantity of 10 showed as "10." in the customer's sheet — so a column of whole
# numbers gets a format with no decimal part at all.
WHOLE_FORMAT = "#,##0"
FRACTION_FORMAT = "#,##0.00"


def quantity_format_for(values: Sequence[Any]) -> str:
    """The format for a column of counts, measured from what is in it.

    Parsed rather than type-checked: a column the reader had no role for arrives
    as the text it printed, and "10" is as whole a number as 10 is.
    """
    numbers = [number for number in (_as_number(value) for value in values)
               if number is not None]
    if numbers and all(float(number).is_integer() for number in numbers):
        return WHOLE_FORMAT
    return FRACTION_FORMAT
_PERCENT_HEADING = re.compile(r"%|percent|rate\s*%|نسبة|بالمئة|بالمائة", re.I)


def percent_column(heading: str, field: str, role: str, values: Sequence[Any]) -> bool:
    """Whether a column of figures is a percentage rather than an amount.

    Asked of the heading first — "ضريبة %" says so outright — and otherwise of
    the numbers, because a tax column holding 0.05 on every row is a rate and
    one holding 250.00 is an amount, and only one of the two should be shown
    with a per-cent sign.
    """
    if _PERCENT_HEADING.search(f"{heading} {field}"):
        return True
    if role not in {"tax", "discount"}:
        return False
    numbers = [value for value in values if isinstance(value, (int, float))
               and not isinstance(value, bool)]
    return bool(numbers) and all(0 < abs(value) < 1 for value in numbers)


# A heading that says the column holds money, when nothing else did.
_MONEY_HEADING = re.compile(
    r"total|amount|value|price|cost|charge|freight|fee"
    r"|إجمالي|المجموع|قيمة|سعر|مبلغ|تكلفة|أجرة",
    re.I,
)


def numeric_text_fields(
    items: Sequence[dict[str, Any]], fields: Sequence[tuple[str, str, str]]
) -> set[str]:
    """Columns the reader returned as text that are really numbers.

    ``qty`` and ``unit_price`` reach the sheet as numbers already, because their
    role is known. A column the role resolver could not name — "Total Value" on
    a manifest, "Gross Weight" — arrives as the text that was printed, and text
    is invisible to SUM and AVERAGE. So a column is measured: if what is in it
    parses as numbers and its heading does not say the digits are an identifier,
    it is written numeric.

    Deliberately by evidence rather than by heading alone. Every document names
    its columns differently, and a heading match would convert an "Order No"
    somebody called "Order Total" on the next customer's paper.
    """
    from table_shape import numeric_cell

    numeric: set[str] = set()
    for field, heading, role in fields:
        if role != "other" or _IDENTIFIER_HEADING.search(f"{field} {heading}"):
            continue
        texts = [
            str(item.get(field)).strip()
            for item in items
            if item.get(field) not in (None, "")
        ]
        if not texts:
            continue
        # A leading zero carries meaning — 007 is a code, and 7 is not it.
        if any(re.match(r"0\d", text) for text in texts):
            continue
        # Strictly "this cell is a number", not "a number appears in it": the
        # second reading turns a column of product codes into a column of
        # meaningless integers.
        parsed = sum(1 for text in texts if numeric_cell(text) is not None)
        if parsed * 5 >= len(texts) * 4:
            numeric.add(field)
    return numeric


ARABIC = re.compile(r"[؀-ۿݐ-ݿﭐ-﷿ﹰ-﻿]")


def reading_order(text: Any) -> int:
    """Excel's per-cell reading order: 0 context, 1 left-to-right, 2 right-to-left.

    Set per cell rather than per sheet because these documents mix scripts
    inside one table — an Arabic description beside an English SKU. A sheet-wide
    setting pushes the punctuation of every foreign cell to the wrong end, so
    each cell is told which way its own text runs.
    """
    value = str(text or "")
    if not value.strip():
        return 0
    arabic = len(ARABIC.findall(value))
    latin = len(re.findall(r"[A-Za-z]", value))
    if arabic and arabic >= latin:
        return 2
    if latin:
        return 1
    return 0


def _text_alignment(text: Any, *, wrap: bool = False, vertical: str = "top"):
    """Align a text cell to the side its own script starts from.

    Left for Latin, right for Arabic. Those are the same rule — text starts at
    the edge its reader starts from — and it is why the alignment is decided per
    cell: an Arabic description sitting beside an English SKU wants a different
    edge from its neighbour. Numbers never come through here; they are right
    aligned against the column of figures above them.
    """
    from openpyxl.styles import Alignment

    order = reading_order(text)
    return Alignment(
        vertical=vertical,
        wrap_text=wrap,
        readingOrder=order,
        horizontal="right" if order == 2 else "left",
    )


def _amount(value: Any) -> float:
    """A recorded total as a number, so the next formula can build on it."""
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        return 0.0
    return float(value)


def _as_number(value: Any) -> float | None:
    """A cell's value as a number, or ``None`` when the cell is not just a number.

    Strict on purpose. The permissive reading — "find a number anywhere in the
    text" — turned a product code column reading "ب ط 001" into the numbers 1,
    2, 4, and the customer lost every product identifier on the invoice.
    """
    from table_shape import numeric_cell

    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    return numeric_cell(value)


def _shown(value: Any) -> str:
    """What a number will look like once Excel has formatted it.

    Measuring the raw ``31.0`` would give a column too narrow for the
    ``31.00 ر.س`` the customer actually sees in it.
    """
    if isinstance(value, bool) or value is None:
        return ""
    if isinstance(value, (int, float)):
        return f"{value:,.2f}"
    return str(value)


def _number_alignment(*, horizontal: str = "right"):
    from openpyxl.styles import Alignment

    return Alignment(horizontal=horizontal, vertical="top")


def _style_cell(cell, *, border, bold: bool = False, size: int = 11) -> None:
    from openpyxl.styles import Font

    cell.font = Font(name=FONT_NAME, size=size, bold=bold)
    cell.border = border


def _track(widths: dict[int, int], column: int, text: Any) -> None:
    """Widen a column to fit what was just written into it.

    openpyxl cannot ask Excel to auto-fit, so the fit is measured here as the
    text is written — which is the only moment every value of a column is known.
    Measured on the longest line rather than the whole string, because a wrapped
    cell breaks at its newlines, and scaled up for Arabic, whose glyphs are
    drawn wider than Latin at the same point size and would otherwise be clipped
    by a column measured in characters.
    """
    value = "" if text is None else str(text)
    longest = max((len(part) for part in value.splitlines()), default=0)
    if ARABIC.search(value):
        longest = int(longest * 1.15) + 1
    widths[column] = max(widths.get(column, MIN_WIDTH), min(longest + 3, MAX_WIDTH))


def _write_page(
    sheet,
    document: dict[str, Any],
    source: Path,
    destination: Path,
    styles: dict[str, Any],
    cache: dict[str, Any] | None = None,
) -> tuple[int, int, list[dict[str, Any]]]:
    """Render one page. Returns (records, flagged rows, review items)."""
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Border, Font, Side

    thin = styles["thin"]
    yellow = styles["yellow"]
    band = styles["band"]
    label_fill = styles["label_fill"]
    stripe = styles.get("stripe")
    currency = money_format(str(document.get("currency") or ""))
    quantity_format = FRACTION_FORMAT

    # One decision, taken once: which way the page reads. It turns the whole
    # sheet around so column A sits where the document's first column sits, and
    # it picks the language of every word this builder adds.
    rtl = str(document.get("direction") or "ltr").casefold() == "rtl"
    sheet.sheet_view.rightToLeft = rtl
    say = document_words(document)

    widths: dict[int, int] = {}
    review_items: list[dict[str, Any]] = []
    records = flagged = 0
    # What each formula works out to, so the cell shows a number even where
    # nothing recalculates it. See :mod:`formula_cache`.
    results: dict[str, Any] = {} if cache is None else cache.setdefault(sheet.title, {})
    line_sum = 0.0

    header = document.get("header") or {}
    header_columns = plan_header_columns(document)
    fields = plan_columns(document)
    items = list(document.get("items") or [])
    numeric_fields = numeric_text_fields(items, fields)
    counts = {
        field: quantity_format_for([item.get(field) for item in items])
        for field, _heading, _role in fields
    }

    # Which page a row came from, kept only when the sheet holds more than one.
    # A merged manifest is a single table, and "page 3" is how a reviewer finds
    # the sheet of paper a disputed row is printed on.
    pages = [page for page in (document.get("pages") or []) if page]
    show_page = len(pages) > 1
    page_column = len(header_columns) + 1 if show_page else 0
    offset = len(header_columns) + (1 if show_page else 0)
    width = max(offset + len(fields), 2)
    row = 1

    def flag(at_row: int, column: int, heading: str, value: Any, note: str) -> None:
        """Queue a cell for manual review.

        Only literal cells are queued. A formula cell must never enter the
        queue: ``apply_review_file`` writes the corrected value straight into
        the cell, which would silently replace the formula with a constant.
        """
        review_items.append({
            "output": str(destination),
            "sheet": sheet.title,
            "row": at_row,
            "column": column,
            "header": heading,
            "value": "" if value is None else value,
            "confidence": "",
            "suggestion": note,
        })

    def heading_cell(column: int, text: str) -> None:
        cell = sheet.cell(row, column, text)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=BAND_TEXT)
        cell.fill = band
        cell.border = thin
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        _track(widths, column, text)

    # ---- one heading row, and then nothing but data ----------------------
    #
    # No document title, no file name, no page banner. The reader lifts whatever
    # is printed largest at the top — "YOUR LOGO", "Shipping Manifest" — and
    # once the pages of one order are a single table, those lines are a heading
    # stranded in the middle of the data. What the file is stays in its name and
    # on the review sheet, which is where a spreadsheet keeps it.
    header_row = row
    for index, (_key, heading) in enumerate(header_columns, start=1):
        heading_cell(index, heading)
    if show_page:
        heading_cell(page_column, say("page_n"))
    role_columns: dict[str, int] = {}
    for index, (_field, heading, role) in enumerate(fields, start=offset + 1):
        heading_cell(index, heading)
        if role != "other" and role not in role_columns:
            role_columns[role] = index
    if header_columns or fields:
        sheet.row_dimensions[row].height = 20
        row += 1

    body_start = row

    def banded(cell, at_row: int) -> None:
        # Banded rows. On a wide grid the eye loses its line between the
        # description and the amount, and a flat table is wide by construction.
        if stripe is not None and (at_row - body_start) % 2 == 1:
            cell.fill = stripe

    def write_header_columns(at_row: int) -> None:
        """The document's own header fields, across one row of the table.

        Repeated on every row on purpose. That is what makes the sheet a table:
        sort it by amount, filter it to one product, paste it under last
        month's, and each line still says who sent it and against which order.
        """
        for index, (key, _heading) in enumerate(header_columns, start=1):
            value = header.get(key)
            text = str(value).strip() if value not in (None, "") else MISSING
            cell = sheet.cell(at_row, index, text)
            _style_cell(cell, border=thin)
            cell.alignment = _text_alignment(text, wrap=len(text) > WRAP_AT)
            banded(cell, at_row)
            _track(widths, index, text)

    # ---- the table -------------------------------------------------------
    first_item_row = last_item_row = 0
    if items and fields:
        first_item_row = row
        qty_column = role_columns.get("qty")
        price_column = role_columns.get("unit_price")
        total_column = role_columns.get("line_total")
        can_compute = bool(qty_column and price_column and total_column)

        for item in items:
            reviews = item.get("review") or {}
            notes = item.get("notes") or {}
            painted = False
            write_header_columns(row)
            if show_page:
                cell = sheet.cell(row, page_column, item.get("_page") or pages[0])
                _style_cell(cell, border=thin)
                cell.alignment = _number_alignment(horizontal="center")
                banded(cell, row)
                _track(widths, page_column, cell.value)

            for index, (field, heading, role) in enumerate(fields, start=offset + 1):
                value = item.get(field)
                # A number is written as a number: the customer's formulas —
                # SUM over a column of totals, AVERAGE over quantities — see
                # nothing at all in a cell holding the text "1,240.00".
                numeric = role in MONEY_ROLES or role in QTY_ROLES
                if field in numeric_fields:
                    converted = _as_number(value)
                    if converted is not None:
                        value, numeric = converted, True

                is_formula = False
                if can_compute and role == "line_total":
                    qty = item.get("qty")
                    price = item.get("unit_price")
                    if qty is not None and price is not None:
                        cell = sheet.cell(
                            row,
                            index,
                            f"={_column_letter(qty_column)}{row}*{_column_letter(price_column)}{row}",
                        )
                        is_formula = True
                        results[cell.coordinate] = qty * price
                        line_sum += qty * price
                        if value is not None:
                            cell.comment = Comment(
                                f"{say('read_value')}: {value:,.2f}", "Vertex"
                            )
                        _track(widths, index, _shown(value if value is not None else qty * price))
                    else:
                        cell = sheet.cell(row, index, value)
                        _track(widths, index, _shown(value))
                elif numeric:
                    # Left genuinely empty when it is missing. "N/A" here would
                    # be text sitting in a column of figures, which is exactly
                    # what makes a total stop adding up.
                    cell = sheet.cell(row, index, value)
                    if role == "line_total" and isinstance(value, (int, float)):
                        line_sum += float(value)
                    _track(widths, index, _shown(value))
                else:
                    text = str(value).strip() if value not in (None, "") else MISSING
                    cell = sheet.cell(row, index, text)
                    _track(widths, index, text)

                _style_cell(cell, border=thin)
                banded(cell, row)
                if role in MONEY_ROLES or (
                    numeric and role == "other" and _MONEY_HEADING.search(f"{field} {heading}")
                ):
                    cell.number_format = currency
                    cell.alignment = _number_alignment()
                elif role in QTY_ROLES:
                    cell.number_format = counts.get(field, quantity_format)
                    cell.alignment = _number_alignment(horizontal="center")
                elif numeric:
                    cell.number_format = counts.get(field, quantity_format)
                    cell.alignment = _number_alignment()
                else:
                    cell.alignment = _text_alignment(
                        cell.value, wrap=len(str(cell.value or "")) > WRAP_AT
                    )

                note = notes.get(field) or notes.get(role)
                if note and not is_formula:
                    cell.comment = Comment(str(note), "Vertex")
                if reviews.get(field) or reviews.get(role):
                    cell.fill = yellow
                    painted = True
                    if not is_formula:
                        flag(row, index, heading, cell.value, str(note or ""))
                    elif qty_column:
                        # Redirect the fix to the quantity, which is a literal.
                        flag(row, qty_column, say("qty"), item.get("qty"), str(note or ""))
            records += 1
            if painted:
                flagged += 1
            row += 1
        last_item_row = row - 1
        sheet.freeze_panes = sheet.cell(first_item_row, 1).coordinate
        # Sort and filter from the headings. A reviewer's first instinct on a
        # flagged invoice is to sort by amount or filter to one product line,
        # and without this they have to select the range by hand every time.
        sheet.auto_filter.ref = (
            f"{_column_letter(1)}{header_row}:{_column_letter(width)}{last_item_row}"
        )
        row += 1
    elif header_columns:
        # A page with fields but no grid — a cover sheet, a delivery note — is
        # still one row of a table rather than a stack of pairs.
        write_header_columns(row)
        sheet.freeze_panes = sheet.cell(row, 1).coordinate
        row += 2

    # ---- totals ----------------------------------------------------------
    totals = document.get("totals") or {}
    totals_review = document.get("totals_review") or {}
    totals_notes = document.get("totals_notes") or {}
    if totals:
        total_column = role_columns.get("line_total")
        # Under the column of amounts it belongs to. Without an amounts column
        # the totals still go at the right-hand end of the item block, never
        # into the header columns on the left, where they would read as another
        # shipper detail.
        value_column = total_column or (
            offset + 2 if len(fields) >= 2 else max(2, offset + 1)
        )
        label_column = max(1, value_column - 1)
        written: dict[str, int] = {}
        totals_results: dict[str, Any] = {}
        for key in _TOTAL_ORDER:
            # The sheet adds the column up only when the page printed no sum of
            # its own. A document that states its total gets that total, not a
            # second one beside it.
            if key not in totals and not (
                key == "subtotal" and first_item_row and "grand_total" not in totals
            ):
                continue
            value = totals.get(key)
            label_cell = sheet.cell(row, label_column, say(key, key))
            is_grand = key == "grand_total"
            _style_cell(label_cell, border=thin, bold=True)
            label_cell.fill = label_fill
            # Pushed up against its own amount. The label sits one column before
            # the value, and a right-to-left sheet mirrors the columns — so the
            # side that puts the two together flips with the sheet.
            label_cell.alignment = Alignment(horizontal="left" if rtl else "right")
            _track(widths, label_column, label_cell.value)

            formula = None
            computed = None
            if key == "subtotal" and total_column and first_item_row and last_item_row:
                letter = _column_letter(total_column)
                formula = f"=SUM({letter}{first_item_row}:{letter}{last_item_row})"
                computed = line_sum
            elif key == "tax_amount" and "subtotal" in written and totals.get("tax_rate"):
                letter = _column_letter(value_column)
                rate = float(totals["tax_rate"])
                formula = f"={letter}{written['subtotal']}*{rate}"
                computed = _amount(totals_results.get("subtotal")) * rate
            elif is_grand and "subtotal" in written and "subtotal" in totals:
                letter = _column_letter(value_column)
                parts = [f"{letter}{written['subtotal']}"]
                computed = _amount(totals_results.get("subtotal"))
                if "tax_amount" in written:
                    parts.append(f"+{letter}{written['tax_amount']}")
                    computed += _amount(totals_results.get("tax_amount"))
                if "discount" in written:
                    parts.append(f"-ABS({letter}{written['discount']})")
                    computed -= abs(_amount(totals_results.get("discount")))
                formula = "=" + "".join(parts)

            value_cell = sheet.cell(row, value_column, formula if formula else value)
            totals_results[key] = computed if computed is not None else value
            if formula is not None and computed is not None:
                results[value_cell.coordinate] = computed
            _style_cell(value_cell, border=thin, bold=True, size=12 if is_grand else 11)
            value_cell.number_format = currency
            value_cell.alignment = Alignment(horizontal="right")
            if formula is not None and value is not None:
                value_cell.comment = Comment(
                    f"{say('read_value')}: {value:,.2f}", "Vertex"
                )
            _track(widths, value_column, f"{value or 0:,.2f}")
            if is_grand:
                value_cell.border = Border(
                    left=thin.left, right=thin.right,
                    top=Side(style="thin", color=BAND),
                    bottom=Side(style="double", color=BAND),
                )
            note = totals_notes.get(key)
            if note and formula is None:
                value_cell.comment = Comment(str(note), "Vertex")
            if totals_review.get(key):
                value_cell.fill = yellow
                flagged += 1
                if formula is None:
                    flag(row, value_column, str(label_cell.value), value, str(note or ""))
            written[key] = row
            row += 1
        row += 1

    # There is deliberately no free-text section here. Everything the reader
    # picked up outside the grid now arrives either as a header column or as a
    # figure in the totals; whatever is left is page furniture — logos, slogans,
    # the repeated document title — and dumping it under the table put a second,
    # unaligned "sheet" inside the one the customer works in.

    if row == 1:
        sheet.cell(1, 1, say("no_page_data"))
    for column, size in widths.items():
        sheet.column_dimensions[_column_letter(column)].width = max(MIN_WIDTH, size)

    # Printable as it stands. These sheets get printed and passed around an
    # accounts office, and a nine-column invoice spilling onto a second sheet
    # with no headings on it is useless on paper.
    sheet.page_setup.orientation = "landscape" if width > 6 else "portrait"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.print_options.horizontalCentered = True
    if first_item_row:
        sheet.print_title_rows = f"{header_row}:{header_row}"
    return records, flagged, review_items


def _data_sheet_title(base: str, page: int, taken: set[str]) -> str:
    """A name for the flat data sheet that no other sheet has taken."""
    candidate, suffix = base[:31], 2
    while candidate in taken:
        candidate = f"{base[:28]}_{suffix}"[:31]
        suffix += 1
    taken.add(candidate)
    return candidate


def _sheet_title(source: Path, page: int, total: int, taken: set[str]) -> str:
    base = re.sub(r"[\\/*?:\[\]]", "-", source.stem)[:26] or "Extracted"
    title = base if total <= 1 else f"{base[:24]}-{page}"
    candidate, suffix = title[:31], 2
    while candidate in taken:
        candidate = f"{title[:28]}_{suffix}"[:31]
        suffix += 1
    taken.add(candidate)
    return candidate


def _write_summary(book, styles, source: Path, documents, records: int,
                   review_items: list[dict[str, Any]]) -> None:
    """A front sheet saying what was read and what still needs a human.

    Put first, and deliberately: the reviewer's question is never "what does row
    41 say" but "is any of this wrong". A workbook that opens on a wall of
    figures makes them hunt for the yellow cells; one that opens on a list of
    them, with the reason beside each, turns the check into a short read.
    """
    from openpyxl.styles import Alignment, Font, PatternFill

    # The front sheet follows the documents behind it. A majority, not "any one
    # of them": a batch of English invoices with a single Arabic page in it is
    # still an English workbook to the person opening it.
    rtl_pages = sum(1 for d in documents if str(d.get("direction") or "") == "rtl")
    rtl = rtl_pages * 2 > len(documents) if documents else False
    say = words_for("rtl" if rtl else "ltr")

    sheet = book.create_sheet(title=say("review"), index=0)
    thin, band, label_fill = styles["thin"], styles["band"], styles["label_fill"]
    sheet.sheet_view.rightToLeft = rtl
    sheet.sheet_view.showGridLines = False

    def band_row(row: int, text: str, span: int = 4) -> int:
        cell = sheet.cell(row, 1, text)
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color=BAND_TEXT)
        cell.fill = band
        cell.border = thin
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
        sheet.row_dimensions[row].height = 20
        return row + 1

    def pair(row: int, label: str, value: Any) -> int:
        left = sheet.cell(row, 1, label)
        left.font = Font(name=FONT_NAME, size=11, bold=True)
        left.fill = label_fill
        left.border = thin
        left.alignment = _text_alignment(label)
        right = sheet.cell(row, 2, value)
        _style_cell(right, border=thin)
        right.alignment = _text_alignment(value)
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        return row + 1

    title = sheet.cell(1, 1, say("review_title"))
    title.font = Font(name=FONT_NAME, size=16, bold=True, color=TITLE_TEXT)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    row = 3

    row = band_row(row, say("source"))
    row = pair(row, say("file"), source.name)
    # Pages read, not sheets written. The two stopped being the same number when
    # the pages of one order started arriving as one table.
    row = pair(row, say("pages"), sum(len(d.get("pages") or [1]) for d in documents))
    row = pair(row, say("items_extracted"), records)
    row += 1

    row = band_row(row, say("what_checked"))
    # Named so the reviewer knows what the absence of a flag actually means.
    row = pair(row, say("shape"), say("shape_note"))
    row = pair(row, say("arithmetic"), say("arithmetic_note"))
    # Only claimed when it happened. The check needs the local reader, and on a
    # server without it the gate passes everything in silence.
    checked = all(d.get("evidence_checked") for d in documents) if documents else False
    row = pair(row, say("pixels"), say("pixels_note") if checked else say("pixels_off"))
    if not checked:
        sheet.cell(row - 1, 2).fill = styles["yellow"]
    # A figure the arithmetic proved was misread and rewrote. Counted here
    # because a correction nobody can see is a correction nobody can audit.
    corrections = sum(len(document.get("repaired") or []) for document in documents)
    if corrections:
        row = pair(row, f"{say('repaired')} ({corrections})", say("repaired_note"))
    row += 1

    row = band_row(row, say("needs_review"))
    if not review_items:
        cell = sheet.cell(row, 1, say("all_clean"))
        _style_cell(cell, border=thin)
        cell.alignment = _text_alignment(cell.value)
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 1
    else:
        for index, heading in enumerate(
                (say("col_sheet"), say("col_cell"), say("col_value"), say("col_reason")), start=1):
            cell = sheet.cell(row, index, heading)
            cell.font = Font(name=FONT_NAME, size=11, bold=True, color=BAND_TEXT)
            cell.fill = band
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
        row += 1
        for item in review_items:
            where = f"{_column_letter(int(item.get('column') or 1))}{item.get('row') or ''}"
            for index, value in enumerate(
                # "suggestion" is the key `flag` writes; it carries the reason
                # the gate objected, which is the only column a reviewer reads.
                (item.get("sheet") or "", where, item.get("value"),
                 item.get("suggestion") or say("needs_confirming")), start=1
            ):
                cell = sheet.cell(row, index, value)
                _style_cell(cell, border=thin)
                cell.alignment = _text_alignment(value, wrap=index == 4)
            row += 1

    for column, width in ((1, 26), (2, 34), (3, 20), (4, 52)):
        sheet.column_dimensions[_column_letter(column)].width = width
    sheet.page_setup.fitToWidth = 1
    sheet.sheet_properties.pageSetUpPr.fitToPage = True


def write_ai_workbook(
    destination: Path,
    source: Path,
    documents: Sequence[dict[str, Any]],
) -> tuple[int, int, list[dict[str, Any]], dict[str, Any], Path]:
    """Build the workbook. Same 5-tuple as the writers in :mod:`export`."""
    from openpyxl import Workbook
    from openpyxl.styles import Border, PatternFill, Side

    book = Workbook()
    book.remove(book.active)
    side = Side(style="thin", color=RULE)
    styles = {
        "thin": Border(left=side, right=side, top=side, bottom=side),
        "yellow": PatternFill(fill_type="solid", fgColor=YELLOW),
        "band": PatternFill(fill_type="solid", fgColor=BAND),
        "label_fill": PatternFill(fill_type="solid", fgColor=LABEL_FILL),
        "stripe": PatternFill(fill_type="solid", fgColor=STRIPE),
    }

    import document_sheet

    # What every formula works out to, gathered as the sheets are written and
    # stored into the file afterwards. Without it a formula cell shows empty
    # anywhere that will not recalculate — Excel's Protected View above all,
    # which is how a downloaded workbook always opens.
    cache: dict[str, dict[str, Any]] = {}
    records = low = 0
    review_items: list[dict[str, Any]] = []
    taken: set[str] = set()
    first_headings: list[str] = []
    rtl_pages = sum(1 for d in documents if str(d.get("direction") or "") == "rtl")
    dominant_direction = "rtl" if documents and rtl_pages * 2 > len(documents) else "ltr"

    # The document first, reproduced as printed. This is the sheet the customer
    # opens and compares against the paper in their hand, so it is the one the
    # review queue points into and the one the workbook opens on.
    for index, document in enumerate(documents, start=1):
        sheet = book.create_sheet(title=_sheet_title(source, index, len(documents), taken))
        page_records, page_low, page_review = document_sheet.write_document(
            sheet, document, source, destination, styles, cache
        )
        records += page_records
        low += page_low
        review_items.extend(page_review)
        if not first_headings:
            first_headings = [heading for _field, heading, _role in plan_columns(document)]

    # Then the same line items as one flat table — every row carrying the
    # document's own header fields — for the customer who wants to sort, filter
    # and pivot rather than read. Only written when there is a table to write.
    for index, document in enumerate(documents, start=1):
        if not document.get("items"):
            continue
        say = document_words(document)
        sheet = book.create_sheet(title=_data_sheet_title(say("data_sheet"), index, taken))
        _write_page(sheet, document, source, destination, styles, cache)

    if not book.worksheets:
        sheet = book.create_sheet(title="Extracted")
        sheet.cell(1, 1, words_for(dominant_direction)("no_file_data"))

    _write_summary(book, styles, source, list(documents), records, review_items)
    # The review summary is added at index 0, so put the document back in front:
    # what the customer wants first is their invoice, not our notes about it.
    book.move_sheet(book.worksheets[0], offset=len(book.worksheets) - 1)
    book.active = 0

    destination = _save_workbook(book, destination)
    # The formulas keep their answers, so the numbers are visible before
    # anything recalculates. See :mod:`formula_cache`.
    from formula_cache import cache_formula_values

    cache_formula_values(destination, cache)
    for item in review_items:
        item["output"] = str(destination)
    kind = str((documents[0] if documents else {}).get("document_type") or "document")
    template = build_template(
        source.name,
        kind if kind in {"invoice", "table"} else "document",
        json.dumps((documents[0] if documents else {}).get("header") or {}, ensure_ascii=False),
        first_headings,
    )
    return records, low, review_items, template, destination
