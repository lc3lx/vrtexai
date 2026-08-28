"""Release hardening: stress, crash isolation, cancel, paths, XLS, templates, temp."""
from __future__ import annotations

import json
import os
import shutil
import subprocess
import sys
import tempfile
import time
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import load_workbook

from templates import load_templates, match_template
from tests.test_production import run_worker


def _csv(folder: Path, name: str, rows: list[list[str]] | None = None) -> Path:
    path = folder / name
    payload = rows or [["name", "qty"], ["item", "1"]]
    path.write_text("\n".join(",".join(row) for row in payload), encoding="utf-8")
    return path


def _child_pids(pid: int) -> list[int]:
    script = f"(Get-CimInstance Win32_Process -Filter 'ParentProcessId={pid}').ProcessId"
    completed = subprocess.run(
        ["powershell", "-NoProfile", "-Command", script],
        capture_output=True, text=True, timeout=20,
    )
    return [int(token) for token in completed.stdout.split() if token.strip().isdigit()]


def _working_set(pid: int) -> tuple[int, int]:
    script = (
        f"$p = Get-Process -Id {pid} -ErrorAction SilentlyContinue; "
        "if ($p) { '{0} {1}' -f $p.WorkingSet64, $p.PeakWorkingSet64 } else { '0 0' }"
    )
    completed = subprocess.run(
        ["powershell", "-NoProfile", "-Command", script],
        capture_output=True, text=True, timeout=20,
    )
    parts = completed.stdout.split()
    if len(parts) >= 2:
        return int(parts[0]), int(parts[1])
    return 0, 0


def _start_worker(inputs: list[Path], output_dir: Path, workers: int = 2) -> tuple[subprocess.Popen, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    request_path = output_dir.parent / f"req-{output_dir.name}.json"
    result_path = output_dir.parent / f"res-{output_dir.name}.json"
    request_path.write_text(json.dumps({
        "inputPaths": [str(path) for path in inputs],
        "outputDirectory": str(output_dir),
        "resultPath": str(result_path),
        "maxWorkers": workers,
    }, ensure_ascii=False), encoding="utf-8")
    env = os.environ.copy()
    env["PYTHONPATH"] = str(ROOT)
    process = subprocess.Popen(
        [sys.executable, str(ROOT / "main.py"), "--request", str(request_path)],
        stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, env=env,
    )
    return process, result_path


class StressTests(unittest.TestCase):
    def _batch(self, count: int) -> dict:
        temp = Path(tempfile.mkdtemp(prefix=f"vertex-stress-{count}-"))
        try:
            files = [_csv(temp, f"f{index:04}.csv") for index in range(count)]
            output = temp / "out"
            started = time.perf_counter()
            result = run_worker(files, output, workers=min(4, os.cpu_count() or 2))
            result["_elapsed_local"] = round(time.perf_counter() - started, 3)
            leftovers = list(output.glob(".*.sqlite")) + list(Path(tempfile.gettempdir()).glob("vertex-office-*"))
            self.assertEqual(leftovers, [], f"temp leak: {leftovers[:5]}")
            for path in output.glob("*_cleaned_*.xlsx"):
                book = load_workbook(path, read_only=True, data_only=True)
                self.assertTrue(book.sheetnames)
                self.assertGreaterEqual(book.active.max_row, 2)
                book.close()
            return result
        finally:
            shutil.rmtree(temp, ignore_errors=True)

    def test_250_files(self):
        result = self._batch(250)
        self.assertEqual(result["filesProcessed"], 250)
        self.assertEqual(result["filesFailed"], 0)
        self.assertEqual(result["recordsExtracted"], 250)

    def test_500_files(self):
        result = self._batch(500)
        self.assertEqual(result["filesProcessed"], 500)
        self.assertEqual(result["filesFailed"], 0)

    def test_1000_files(self):
        result = self._batch(1000)
        self.assertEqual(result["filesProcessed"], 1000)
        self.assertEqual(result["filesFailed"], 0)
        self.assertEqual(result["recordsExtracted"], 1000)


class CrashAndCancelTests(unittest.TestCase):
    def test_kill_one_child_does_not_kill_batch(self):
        temp = Path(tempfile.mkdtemp(prefix="vertex-crash-"))
        try:
            files = [_csv(temp, f"c{index:03}.csv", [["n"], [str(index)]]) for index in range(24)]
            process, result_path = _start_worker(files, temp / "out", workers=2)
            time.sleep(1.2)
            children = _child_pids(process.pid)
            if children:
                subprocess.run(["taskkill", "/F", "/PID", str(children[0])], capture_output=True)
            stdout, stderr = process.communicate(timeout=120)
            self.assertIsNotNone(process.returncode)
            if result_path.is_file():
                payload = json.loads(result_path.read_text(encoding="utf-8-sig"))
                self.assertGreaterEqual(payload["filesProcessed"], 1)
                self.assertLess(payload["filesFailed"], payload["filesProcessed"])
            else:
                self.assertNotEqual(process.returncode, 0)
            leftovers = _child_pids(process.pid)
            self.assertEqual(leftovers, [])
        finally:
            shutil.rmtree(temp, ignore_errors=True)

    def test_cancel_parent_then_new_batch_works(self):
        temp = Path(tempfile.mkdtemp(prefix="vertex-cancel-"))
        try:
            files = [_csv(temp, f"k{index:03}.csv") for index in range(40)]
            process, _result_path = _start_worker(files, temp / "out", workers=3)
            time.sleep(0.8)
            subprocess.run(["taskkill", "/F", "/T", "/PID", str(process.pid)], capture_output=True)
            stdout, stderr = process.communicate(timeout=20)
            time.sleep(0.4)
            orphans = _child_pids(process.pid)
            self.assertEqual(orphans, [])
            second = run_worker([_csv(temp, "after.csv")], temp / "out2", workers=1)
            self.assertEqual(second["filesFailed"], 0)
            self.assertGreaterEqual(second["recordsExtracted"], 1)
        finally:
            shutil.rmtree(temp, ignore_errors=True)


class PathAndXlsTests(unittest.TestCase):
    def test_unicode_spaces_duplicates_and_missing_output_dir(self):
        temp = Path(tempfile.mkdtemp(prefix="vertex-path-"))
        try:
            first = _csv(temp, "فاتورة عميل.csv", [["الاسم", "الكمية"], ["علي", "2"]])
            second = _csv(temp, "my file.csv", [["name", "qty"], ["A", "1"]])
            nested = temp / "a" / "dup.csv"
            other = temp / "b" / "dup.csv"
            nested.parent.mkdir()
            other.parent.mkdir()
            _csv(nested.parent, "dup.csv", [["x"], ["1"]])
            _csv(other.parent, "dup.csv", [["x"], ["2"]])
            missing = temp / "not-created-yet" / "out"
            result = run_worker([first, second, nested, other], missing, workers=2)
            self.assertEqual(result["filesProcessed"], 4)
            self.assertEqual(result["filesFailed"], 0)
            cleaned = list(missing.glob("*_cleaned_*.xlsx"))
            self.assertEqual(len(cleaned), 4)
            book = load_workbook(next(p for p in cleaned if "فاتورة" in p.name), read_only=True, data_only=True)
            self.assertEqual(next(book.active.iter_rows(values_only=True))[0], "الاسم")
            book.close()
        finally:
            shutil.rmtree(temp, ignore_errors=True)

    def test_real_xls_via_excel_or_skip(self):
        temp = Path(tempfile.mkdtemp(prefix="vertex-xls-"))
        xls = temp / "real.xls"
        script = (
            f"$excel = New-Object -ComObject Excel.Application; $excel.Visible = $false; $excel.DisplayAlerts = $false; "
            f"$wb = $excel.Workbooks.Add(); $ws = $wb.Worksheets.Item(1); "
            f"$ws.Cells.Item(1,1) = 'الاسم'; $ws.Cells.Item(1,2) = 'التاريخ'; $ws.Cells.Item(1,3) = 'القيمة'; "
            f"$ws.Cells.Item(2,1) = 'شركة الاختبار'; $ws.Cells.Item(2,2) = '2024-03-01'; $ws.Cells.Item(2,3) = 1250.5; "
            f"$ws2 = $wb.Worksheets.Add(); $ws2.Name = 'Second'; $ws2.Cells.Item(1,1) = 'code'; $ws2.Cells.Item(2,1) = 'O12'; "
            f"$wb.SaveAs('{str(xls).replace(chr(39), chr(39)+chr(39))}', 56); $wb.Close($false); $excel.Quit()"
        )
        created = subprocess.run(["powershell", "-NoProfile", "-Command", script], capture_output=True, text=True, timeout=60)
        try:
            if created.returncode != 0 or not xls.is_file():
                self.skipTest(f"Excel COM unavailable: {(created.stderr or created.stdout)[:200]}")
            result = run_worker([xls], temp / "out", workers=1)
            self.assertEqual(result["filesFailed"], 0)
            self.assertGreaterEqual(result["recordsExtracted"], 1)
            out = next((temp / "out").glob("*_cleaned_*.xlsx"))
            book = load_workbook(out, data_only=True)
            self.assertGreaterEqual(len(book.sheetnames), 2)
            texts = []
            for sheet in book.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    texts.extend(str(cell) for cell in row if cell is not None)
            book.close()
            self.assertIn("الاسم", texts)
            self.assertIn("شركة الاختبار", texts)
            self.assertTrue(any(value in {"O12", "012"} for value in texts))
        finally:
            shutil.rmtree(temp, ignore_errors=True)


class TemplateRecoveryTests(unittest.TestCase):
    def test_broken_template_shapes_do_not_crash(self):
        temp = Path(tempfile.mkdtemp(prefix="vertex-tpl2-"))
        try:
            (temp / "templates.json").write_text("[]", encoding="utf-8")
            self.assertEqual(load_templates(temp / "templates.json"), [])
            (temp / "templates.json").write_text("{}", encoding="utf-8")
            self.assertEqual(load_templates(temp / "templates.json"), [])
            (temp / "templates.json").write_text('["x", {"type": "nope", "fingerprint": "z"}]', encoding="utf-8")
            loaded = load_templates(temp / "templates.json")
            self.assertIsNone(match_template("anything at all", loaded))
            (temp / "templates.json").write_text(
                '[{"type": "invoice", "header_text": "Hello", "regions": {"bad": [-1, None]}}]',
                encoding="utf-8",
            )
            loaded = load_templates(temp / "templates.json")
            match_template("Hello world", loaded)
            csv = _csv(temp, "ok.csv")
            result = run_worker([csv], temp / "out", workers=1)
            self.assertEqual(result["filesFailed"], 0)
        finally:
            shutil.rmtree(temp, ignore_errors=True)


class IntegrityAndTempTests(unittest.TestCase):
    def test_output_reopen_and_temp_clean(self):
        temp = Path(tempfile.mkdtemp(prefix="vertex-int-"))
        before_office = set(Path(tempfile.gettempdir()).glob("vertex-office-*"))
        try:
            testdata = ROOT / "testdata"
            result = run_worker(
                [testdata / "Complex_Table_Report.xlsx", testdata / "Comprehensive_OCR_Test_Data.xlsx"],
                temp / "out", workers=2,
            )
            self.assertEqual(result["recordsExtracted"], 5500)
            for path in (temp / "out").glob("*_cleaned_*.xlsx"):
                book = load_workbook(path, read_only=True, data_only=True)
                self.assertTrue(book.sheetnames)
                row = next(book.active.iter_rows(values_only=True))
                self.assertTrue(any(cell for cell in row))
                book.close()
            after_office = set(Path(tempfile.gettempdir()).glob("vertex-office-*"))
            self.assertEqual(after_office - before_office, set())
            self.assertEqual(list((temp / "out").glob(".*.sqlite")), [])
        finally:
            shutil.rmtree(temp, ignore_errors=True)

    def test_mixed_success_failure_isolation(self):
        temp = Path(tempfile.mkdtemp(prefix="vertex-iso-"))
        try:
            good_a = _csv(temp, "a.csv")
            bad = temp / "b.pdf"
            bad.write_bytes(b"not-a-pdf")
            good_c = _csv(temp, "c.csv", [["h"], ["z"]])
            warn = ROOT / "testdata" / "Complex_Table_Report.xlsx"
            result = run_worker([good_a, bad, good_c, warn], temp / "out", workers=2)
            self.assertEqual(result["filesProcessed"], 4)
            self.assertGreaterEqual(result["filesFailed"], 1)
            self.assertGreaterEqual(result["filesSucceeded"] + result["filesWithWarnings"], 2)
        finally:
            shutil.rmtree(temp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
