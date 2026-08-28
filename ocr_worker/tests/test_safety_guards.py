"""Regression tests for conservative OCR-to-Excel decisions."""
from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from clean import find_reliable_header_row
from export import write_generic_tables
from invoice import invoice_table_is_reliable, parse_invoice_table
from main import _process_extracted
from ocr import _postprocess_table


class SafetyGuardTests(unittest.TestCase):
    def test_ocr_record_cannot_replace_a_missing_header(self):
        rows = [
            ["4765", "شركة مثال للمقاولات", "user@example.com 966555000000", "مقاول"],
            ["ID", "Company", "Email", "Category"],
            ["4766", "شركة ثانية", "second@example.com", "مقاول"],
        ]
        self.assertEqual(find_reliable_header_row(rows), 1)

    def test_merged_invoice_cells_fall_back_from_semantic_invoice_export(self):
        rows = [
            ["Description", "Quantity", "Rate", "Total"],
            ["Fridge Air Conditioner", "12 Nos 11 Nos", "993.00 1197.00", "11916.00 13167.00"],
        ]
        self.assertFalse(invoice_table_is_reliable(rows))

    def test_clean_invoice_grid_is_allowed(self):
        rows = [
            ["Description", "SKU", "Qty", "Unit Price", "Total"],
            ["Fridge", "FR-1", "2", "10.00", "20.00"],
        ]
        self.assertTrue(invoice_table_is_reliable(rows))

    def test_visible_ocr_noise_in_an_item_description_stays_reviewable(self):
        parsed = parse_invoice_table(
            [["Description", "Qty", "Unit Price", "Total"], ["خدمة ©", "1", "10.00", "10.00"]],
            [[95.0, 95.0, 95.0, 95.0], [95.0, 95.0, 95.0, 95.0]],
        )
        self.assertTrue(parsed["items"][0]["review"])

    def test_postprocess_never_guesses_a_missing_decimal_point(self):
        table, _scores = _postprocess_table(
            [["STATE", "P_CAP"], ["ARIZONA", "1159826"]], [[95.0, 95.0], [95.0, 40.0]]
        )
        self.assertEqual(table[1][1], "1159826")

    def test_generic_review_highlights_only_flagged_cell(self):
        from openpyxl import load_workbook

        temp = Path(tempfile.mkdtemp(prefix="vertex-safety-"))
        destination = temp / "out.xlsx"
        try:
            _records, low, review, _template, _path = write_generic_tables(
                destination,
                Path("input.png"),
                [{
                    "name": "Extracted",
                    "headers": ["Name", "Amount"],
                    "rows": [{
                        "values": ["Alpha", "1159826"],
                        "confidences": [95.0, 40.0],
                        "review": True,
                        "review_columns": [1],
                    }],
                }],
            )
            self.assertEqual(low, 1)
            self.assertEqual(len(review), 1)
            book = load_workbook(destination)
            sheet = book["Extracted"]
            self.assertEqual(sheet["A2"].fill.fill_type, None)
            self.assertEqual(sheet["B2"].fill.fill_type, "solid")
            book.close()
        finally:
            shutil.rmtree(temp, ignore_errors=True)

    def test_unreliable_invoice_keeps_raw_page_context(self):
        from openpyxl import load_workbook

        temp = Path(tempfile.mkdtemp(prefix="vertex-safety-"))
        try:
            result = _process_extracted(
                Path("invoice.png"),
                [[
                    ["Tax Invoice"],
                    ["Description", "Quantity", "Rate", "Total"],
                    ["Fridge Air Conditioner", "12 Nos 11 Nos", "993.00 1197.00", "11916.00 13167.00"],
                ]],
                [[[95.0], [95.0, 95.0, 95.0, 95.0], [50.0, 50.0, 50.0, 50.0]]],
                ["Supplier: Gulf Enterprises\nInvoice No: 118\nTotal: 81,676.00"],
                {},
                temp,
                [],
            )
            book = load_workbook(result.output, read_only=True, data_only=True)
            self.assertIn("OCR Context", book.sheetnames)
            values = [row[0] for row in book["OCR Context"].iter_rows(min_row=2, values_only=True)]
            self.assertIn("Invoice No: 118", values)
            book.close()
        finally:
            shutil.rmtree(temp, ignore_errors=True)

    def test_local_form_schema_is_not_reinterpreted_as_a_later_business_row(self):
        from openpyxl import load_workbook

        temp = Path(tempfile.mkdtemp(prefix="vertex-safety-"))
        try:
            result = _process_extracted(
                Path("rental-invoice.png"),
                [[
                    ["Field", "Value"],
                    ["Mobile", "0501234001"],
                    ["Car Number", "10001"],
                    ["Car Name", "KIA"],
                ]],
                [[[100.0, 100.0], [95.0, 95.0], [95.0, 95.0], [95.0, 95.0]]],
                [],
                {},
                temp,
                [],
                force_generic=True,
                local_form=True,
            )
            book = load_workbook(result.output, read_only=True, data_only=True)
            self.assertEqual(book.sheetnames, ["Extracted"])
            values = list(book["Extracted"].iter_rows(values_only=True))
            self.assertEqual(values[0], ("Field", "Value"))
            self.assertEqual(values[1], ("Mobile", "0501234001"))
            self.assertEqual(values[3], ("Car Name", "KIA"))
            book.close()
        finally:
            shutil.rmtree(temp, ignore_errors=True)


if __name__ == "__main__":
    unittest.main()
