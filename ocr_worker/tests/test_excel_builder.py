"""Tests for the formatted, formula-bearing workbook.

Two things are pinned down here.

:class:`FormulaIntegrityTests` is the important one. openpyxl writes formulas
without evaluating them, so a broken reference is invisible until the customer
opens the file and Excel shows ``#REF!`` or ``#VALUE!``. Rather than install an
office suite to recalculate, every formula is parsed here and each cell it
points at is checked for existence and for holding a number — which is exactly
the condition under which those two errors appear.

:class:`FlatTableTests` and :class:`DataTypeTests` pin the shape the customer
asked for: the document header as columns beside the line items rather than a
stack of pairs above them, no page-text section, no repeated document title,
numbers stored as numbers, and an empty field that keeps its place in the row
instead of shifting everything after it.

Because the header now occupies the first columns, no test may assume the item
table starts at column A. Every column is found by its heading, which is also
how the customer finds it.
"""
from __future__ import annotations

import re
import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import excel_builder

CELL_REFERENCE = re.compile(r"(?<![A-Z0-9_!:])(\$?[A-Z]{1,3}\$?\d{1,7})(?![(\d])")
RANGE_REFERENCE = re.compile(r"(\$?[A-Z]{1,3}\$?\d{1,7}):(\$?[A-Z]{1,3}\$?\d{1,7})")
ERROR_LITERALS = ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#NULL!", "#NUM!")


def document(**overrides) -> dict:
    base = {
        "document_type": "invoice",
        "direction": "rtl",
        "currency": "SAR",
        "title": "فاتورة ضريبية",
        "page": 1,
        "pages": [1],
        "header": {"supplier": "شركة الأفق", "invoice_number": "INV-2201"},
        "columns": ["الوصف", "الكمية", "سعر الوحدة", "الإجمالي"],
        "column_roles": ["description", "qty", "unit_price", "line_total"],
        "items": [
            {"description": "قلم", "qty": 2, "unit_price": 15.5, "line_total": 31.0,
             "review": {}, "notes": {}},
            {"description": "دفتر", "qty": 3, "unit_price": 10.0, "line_total": 30.0,
             "review": {}, "notes": {}},
            {"description": "ممحاة", "qty": 10, "unit_price": 1.25, "line_total": 12.5,
             "review": {}, "notes": {}},
        ],
        "totals": {"subtotal": 73.5, "tax_rate": 0.15, "tax_amount": 11.03, "grand_total": 84.53},
        "totals_review": {},
        "totals_notes": {},
    }
    base.update(overrides)
    return base


class WorkbookCase(unittest.TestCase):
    """Builds a workbook once per test and reopens it the way Excel would."""

    def build(self, *documents):
        from openpyxl import load_workbook

        directory = tempfile.TemporaryDirectory()
        self.addCleanup(directory.cleanup)
        source = Path(directory.name) / "invoice.png"
        destination = Path(directory.name) / "invoice_cleaned.xlsx"
        result = excel_builder.write_ai_workbook(
            destination, source, list(documents) or [document()]
        )
        self.result = result
        self.destination = result[4]
        self.book = load_workbook(self.destination)
        self.addCleanup(self.book.close)
        # Three kinds of sheet, in the order the customer meets them: the page
        # reproduced as printed, then the same line items as a flat table, then
        # the review summary. Found by the names the builder gives them in the
        # document's own language, so a sheet added later cannot silently point
        # every assertion somewhere else.
        say = excel_builder.words_for(
            str((list(documents) or [document()])[0].get("direction") or "ltr")
        )
        self.summary = self.book[say("review")]
        self.assertIs(self.summary, self.book.worksheets[-1])
        data_title = say("data_sheet")
        self.data = self.book[data_title] if data_title in self.book.sheetnames else None
        self.pages = [
            worksheet for worksheet in self.book.worksheets
            if worksheet is not self.summary
            and not worksheet.title.startswith(data_title[:28])
        ]
        # ``document`` is the page as printed; ``sheet`` is the flat data table,
        # which is what most of these assertions are about.
        self.document = self.pages[0]
        self.assertIs(self.document, self.book.worksheets[0])
        self.sheet = self.data if self.data is not None else self.document
        return self.sheet

    # The heading row is row 1: no title, no file name, no page banner above it.
    HEADING_ROW = 1

    def column(self, heading: str, sheet=None) -> int:
        """The column a heading sits in — never a hard-coded letter.

        Searched over the whole sheet rather than the first row: the page sheet
        reproduces whatever the document printed above its table, so the item
        headings are wherever the page put them.
        """
        sheet = sheet if sheet is not None else self.sheet
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == heading:
                    return cell.column
        raise AssertionError(f"لا عمود بعنوان «{heading}»")

    def item_row(self, number: int) -> int:
        """The row the nth item sits on in the page sheet.

        Found rather than counted: the page sheet reproduces whatever blocks the
        document had above its table, so the table does not start at a fixed row.
        """
        return self.find_row("الوصف", self.document) + number

    def find_row(self, text: str, sheet=None) -> int:
        for row in (sheet if sheet is not None else self.sheet).iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == text:
                    return cell.row
        raise AssertionError(f"لم يُعثر على الصف «{text}»")

    def text_of(self, sheet=None) -> str:
        sheet = sheet if sheet is not None else self.sheet
        return "\n".join(
            str(cell.value) for row in sheet.iter_rows() for cell in row
            if isinstance(cell.value, str)
        )

    def formula_cells(self):
        for row in self.sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    yield cell


class FlatTableTests(WorkbookCase):
    """The header belongs beside the items, not stacked above them."""

    def test_header_fields_are_columns_in_the_heading_row(self):
        self.build()
        self.assertEqual(self.column("المورد"), 1)
        self.assertEqual(self.column("رقم الفاتورة"), 2)
        self.assertEqual(self.column("الوصف"), 3)

    def test_a_header_value_repeats_on_every_item_row(self):
        sheet = self.build()
        supplier = self.column("المورد")
        for offset in range(3):
            self.assertEqual(sheet.cell(2 + offset, supplier).value, "شركة الأفق")

    def test_the_header_is_never_a_stack_of_pairs_in_column_a(self):
        # The complaint itself: shipper and consignee arriving as vertical
        # label/value rows in A and B, above a table they had nothing to do with.
        sheet = self.build()
        labels = [sheet.cell(row, 1).value for row in range(2, 6)]
        self.assertNotIn("رقم الفاتورة", labels)

    def test_the_document_title_is_not_written_anywhere(self):
        # "YOUR LOGO" and "Shipping Manifest" are what this lifts in practice.
        self.build(document(title="Shipping Manifest"))
        self.assertNotIn("Shipping Manifest", self.text_of())
        self.assertNotIn("فاتورة ضريبية", self.text_of())

    def test_no_page_text_section_is_written(self):
        # Even when a caller still hands the builder the old key.
        self.build(document(notes=["الدفع خلال 30 يوماً", "YOUR LOGO"]))
        page = self.text_of()
        self.assertNotIn("الدفع خلال 30 يوماً", page)
        self.assertNotIn("YOUR LOGO", page)
        self.assertNotIn("ملاحظات", page)

    def test_a_document_with_no_items_still_writes_its_fields(self):
        # No line items means no data sheet, but the page still has to appear.
        self.build(document(items=[], totals={}))
        self.assertIsNone(self.data)
        self.assertIn("شركة الأفق", self.text_of(self.document))

    def test_a_merged_document_says_which_page_each_row_came_from(self):
        merged = document(pages=[1, 2], items=[
            {"description": "قلم", "qty": 2, "unit_price": 15.5, "line_total": 31.0,
             "_page": 1, "review": {}, "notes": {}},
            {"description": "دفتر", "qty": 3, "unit_price": 10.0, "line_total": 30.0,
             "_page": 2, "review": {}, "notes": {}},
        ], totals={})
        sheet = self.build(merged)
        page = self.column("صفحة")
        self.assertEqual(sheet.cell(2, page).value, 1)
        self.assertEqual(sheet.cell(3, page).value, 2)

    def test_a_single_page_document_has_no_page_column(self):
        self.build()
        with self.assertRaises(AssertionError):
            self.column("صفحة")

    def test_the_page_marker_is_never_mistaken_for_a_data_column(self):
        merged = document(pages=[1, 2])
        merged["items"][0]["_page"] = 1
        self.build(merged)
        self.assertNotIn("_page", self.text_of())

    def test_the_summary_counts_pages_read_not_sheets_written(self):
        self.build(document(pages=[1, 2, 3]))
        values = [cell.value for row in self.summary.iter_rows() for cell in row]
        self.assertIn(3, values)


class DataTypeTests(WorkbookCase):
    """Numbers as numbers, missing values in their place, nothing shifted."""

    def test_quantity_and_price_are_stored_as_numbers(self):
        sheet = self.build()
        self.assertIsInstance(sheet.cell(2, self.column("الكمية")).value, (int, float))
        self.assertIsInstance(sheet.cell(2, self.column("سعر الوحدة")).value, (int, float))

    def test_a_numeric_text_column_is_written_as_numbers(self):
        # "Total Value" on a manifest: no role, so it arrives as printed text,
        # and text is invisible to SUM.
        sheet = self.build(document(
            columns=["الوصف", "Total Value"],
            column_roles=["description", "other"],
            items=[
                {"description": "قلم", "Total Value": "1,240.00", "review": {}, "notes": {}},
                {"description": "دفتر", "Total Value": "310.50", "review": {}, "notes": {}},
            ],
            totals={},
        ))
        cell = sheet.cell(2, self.column("Total Value"))
        self.assertEqual(cell.value, 1240.0)
        self.assertEqual(cell.alignment.horizontal, "right")

    def test_an_identifier_column_stays_text(self):
        # A tracking number turned into 1.02442E+11 is a defect, not a number.
        sheet = self.build(document(
            columns=["الوصف", "AWB Number"],
            column_roles=["description", "other"],
            items=[
                {"description": "قلم", "AWB Number": "102441700123", "review": {}, "notes": {}},
                {"description": "دفتر", "AWB Number": "102441700124", "review": {}, "notes": {}},
            ],
            totals={},
        ))
        self.assertEqual(sheet.cell(2, self.column("AWB Number")).value, "102441700123")

    def test_a_leading_zero_is_never_turned_into_a_number(self):
        sheet = self.build(document(
            columns=["الوصف", "Lot"],
            column_roles=["description", "other"],
            items=[
                {"description": "قلم", "Lot": "007", "review": {}, "notes": {}},
                {"description": "دفتر", "Lot": "012", "review": {}, "notes": {}},
            ],
            totals={},
        ))
        self.assertEqual(sheet.cell(2, self.column("Lot")).value, "007")

    def test_a_missing_text_field_is_written_rather_than_skipped(self):
        sheet = self.build(document(
            header={"supplier": "شركة الأفق", "shipper_phone": ""},
            items=[{"description": "قلم", "qty": 2, "unit_price": 15.5,
                    "line_total": 31.0, "review": {}, "notes": {}}],
            totals={},
        ))
        # The empty field keeps its column, so nothing after it shifts left.
        self.assertEqual(sheet.cell(1, 1).value, "المورد")
        self.assertEqual(sheet.cell(2, 1).value, "شركة الأفق")

    def test_a_missing_item_text_cell_says_na(self):
        sheet = self.build(document(
            columns=["الوصف", "Remarks"],
            column_roles=["description", "other"],
            items=[
                {"description": "قلم", "Remarks": "urgent", "review": {}, "notes": {}},
                {"description": "دفتر", "Remarks": "", "review": {}, "notes": {}},
            ],
            totals={},
        ))
        self.assertEqual(sheet.cell(3, self.column("Remarks")).value, excel_builder.MISSING)

    def test_a_missing_number_is_left_empty_not_filled_with_na(self):
        # "N/A" in a money column is text in a column of figures, which is what
        # stops SUM and AVERAGE being trustworthy.
        sheet = self.build(document(items=[
            {"description": "خدمة", "qty": None, "unit_price": None,
             "line_total": 500.0, "review": {}, "notes": {}},
        ], totals={}))
        self.assertIsNone(sheet.cell(2, self.column("الكمية")).value)

    def test_numbers_are_right_aligned_and_text_is_not_centred(self):
        sheet = self.build()
        first = 2
        self.assertEqual(sheet.cell(first, self.column("سعر الوحدة")).alignment.horizontal, "right")
        self.assertIn(
            sheet.cell(first, self.column("الوصف")).alignment.horizontal, ("left", "right")
        )

    def test_every_column_is_wide_enough_for_what_is_in_it(self):
        from openpyxl.utils import get_column_letter

        long_text = "وصف طويل جداً لصنف يحمل اسماً كاملاً ومواصفات وأبعاداً ورقم دفعة الإنتاج"
        sheet = self.build(document(items=[
            {"description": long_text, "qty": 2, "unit_price": 15.5,
             "line_total": 31.0, "review": {}, "notes": {}},
        ], totals={}))
        description = self.column("الوصف")
        width = sheet.column_dimensions[get_column_letter(description)].width
        self.assertGreaterEqual(width, min(len(long_text), excel_builder.MAX_WIDTH))
        # What will not fit on one line wraps, so nothing is hidden.
        self.assertTrue(sheet.cell(2, description).alignment.wrap_text)


class FormulaTests(WorkbookCase):
    def test_line_total_is_a_product_formula_not_a_number(self):
        sheet = self.build()
        from openpyxl.utils import get_column_letter

        qty = get_column_letter(self.column("الكمية"))
        price = get_column_letter(self.column("سعر الوحدة"))
        total = self.column("الإجمالي")
        self.assertEqual(sheet.cell(2, total).value, f"={qty}2*{price}2")
        self.assertEqual(sheet.cell(4, total).value, f"={qty}4*{price}4")

    def test_subtotal_sums_exactly_the_item_rows(self):
        sheet = self.build()
        from openpyxl.utils import get_column_letter

        total = get_column_letter(self.column("الإجمالي"))
        subtotal_row = self.find_row("المجموع الفرعي")
        self.assertEqual(
            sheet.cell(subtotal_row, self.column("الإجمالي")).value,
            f"=SUM({total}2:{total}4)",
        )

    def test_tax_multiplies_the_subtotal_cell_by_the_rate(self):
        sheet = self.build()
        from openpyxl.utils import get_column_letter

        column = self.column("الإجمالي")
        letter = get_column_letter(column)
        subtotal_row = self.find_row("المجموع الفرعي")
        tax_row = self.find_row("الضريبة")
        self.assertEqual(sheet.cell(tax_row, column).value, f"={letter}{subtotal_row}*0.15")

    def test_grand_total_adds_the_subtotal_and_tax_cells(self):
        sheet = self.build()
        from openpyxl.utils import get_column_letter

        column = self.column("الإجمالي")
        letter = get_column_letter(column)
        subtotal_row = self.find_row("المجموع الفرعي")
        tax_row = self.find_row("الضريبة")
        grand_row = self.find_row("الإجمالي النهائي")
        self.assertEqual(
            sheet.cell(grand_row, column).value,
            f"={letter}{subtotal_row}+{letter}{tax_row}",
        )

    def test_discount_is_subtracted_from_the_grand_total(self):
        sheet = self.build(document(totals={
            "subtotal": 73.5, "discount": 10.0, "tax_amount": 9.53, "grand_total": 73.03,
        }))
        grand_row = self.find_row("الإجمالي النهائي")
        self.assertIn("-ABS(", str(sheet.cell(grand_row, self.column("الإجمالي")).value))

    def test_the_read_value_survives_as_a_comment_on_the_formula(self):
        sheet = self.build()
        comment = sheet.cell(2, self.column("الإجمالي")).comment
        self.assertIsNotNone(comment)
        self.assertIn("31.00", comment.text)

    def test_a_row_missing_a_price_keeps_the_read_total_instead_of_a_formula(self):
        sheet = self.build(document(items=[
            {"description": "خدمة", "qty": None, "unit_price": None, "line_total": 500.0,
             "review": {}, "notes": {}},
        ]))
        self.assertEqual(sheet.cell(2, self.column("الإجمالي")).value, 500.0)


class FormulaIntegrityTests(WorkbookCase):
    """The static replacement for recalculating in an office suite."""

    def _assert_sound(self, sheet):
        checked = 0
        for cell in self.formula_cells():
            formula = cell.value
            targets: list[str] = []
            for start, end in RANGE_REFERENCE.findall(formula):
                targets.extend(self._expand(sheet, start, end))
            trimmed = RANGE_REFERENCE.sub("", formula)
            targets.extend(CELL_REFERENCE.findall(trimmed))
            self.assertTrue(targets, f"صيغة بلا مراجع: {formula}")
            for reference in targets:
                target = sheet[reference.replace("$", "")]
                self.assertIsNotNone(
                    target.value,
                    f"{cell.coordinate} = {formula} يشير إلى خلية فارغة {reference} (#VALUE!)",
                )
                # A formula may reference another formula; both resolve to numbers.
                if isinstance(target.value, str) and target.value.startswith("="):
                    continue
                self.assertIsInstance(
                    target.value, (int, float),
                    f"{cell.coordinate} = {formula} يشير إلى نص في {reference} (#VALUE!)",
                )
                checked += 1
        self.assertGreater(checked, 0, "لم تُكتب أي صيغة في الملف")

    @staticmethod
    def _expand(sheet, start: str, end: str) -> list[str]:
        from openpyxl.utils import range_boundaries, get_column_letter

        left, top, right, bottom = range_boundaries(f"{start}:{end}".replace("$", ""))
        return [
            f"{get_column_letter(column)}{row}"
            for row in range(top, bottom + 1)
            for column in range(left, right + 1)
        ]

    def test_every_reference_resolves_to_a_number(self):
        sheet = self.build()
        self._assert_sound(sheet)

    def test_integrity_holds_for_a_single_item_document(self):
        sheet = self.build(document(items=[
            {"description": "قلم", "qty": 1, "unit_price": 9.0, "line_total": 9.0,
             "review": {}, "notes": {}},
        ], totals={"subtotal": 9.0, "grand_total": 9.0}))
        self._assert_sound(sheet)

    def test_integrity_holds_when_the_header_pushes_the_columns_across(self):
        # Twelve header columns before the first item column: every formula is
        # written from a measured index, so none of them may drift.
        sheet = self.build(document(header={
            f"field_{number}": f"value {number}" for number in range(12)
        }))
        self._assert_sound(sheet)

    def test_no_error_literal_is_ever_written(self):
        self.build()
        for row in self.sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    for literal in ERROR_LITERALS:
                        self.assertNotIn(literal, cell.value)

    def test_no_formula_is_written_when_there_are_no_items(self):
        # An empty table must not leave a SUM over a range that does not exist.
        sheet = self.build(document(items=[], totals={"grand_total": 12.0}))
        for cell in self.formula_cells():
            self.assertNotIn("SUM(", str(cell.value))
        self._ = sheet


class FormattingTests(WorkbookCase):
    def test_every_cell_uses_arial(self):
        sheet = self.build()
        named = [
            cell for row in sheet.iter_rows() for cell in row
            if cell.value is not None and cell.font and cell.font.name
        ]
        self.assertTrue(named)
        for cell in named:
            self.assertEqual(cell.font.name, "Arial", cell.coordinate)

    def test_the_heading_row_is_bold_white_on_the_brand_colour(self):
        sheet = self.build()
        for heading in ("المورد", "الوصف"):
            cell = sheet.cell(1, self.column(heading))
            self.assertTrue(cell.font.bold)
            self.assertIn("FFFFFF", str(cell.font.color.rgb))
            self.assertIn(excel_builder.BAND, str(cell.fill.fgColor.rgb))

    def test_every_item_cell_has_a_thin_border(self):
        sheet = self.build()
        for column in range(1, 7):
            border = sheet.cell(2, column).border
            for side in (border.left, border.right, border.top, border.bottom):
                self.assertEqual(side.style, "thin")

    def test_money_columns_carry_the_document_currency(self):
        sheet = self.build()
        self.assertIn("ر.س", sheet.cell(2, self.column("سعر الوحدة")).number_format)
        self.assertIn("#,##0.00", sheet.cell(2, self.column("الإجمالي")).number_format)

    def test_a_whole_quantity_is_shown_without_a_decimal_point(self):
        # Excel renders 10 as "10." under a format that allows decimals, which
        # is what the customer saw down their quantity column.
        sheet = self.build()
        self.assertEqual(sheet.cell(2, self.column("الكمية")).number_format, "#,##0")

    def test_a_fractional_quantity_keeps_its_decimals(self):
        sheet = self.build(document(items=[
            {"description": "قماش", "qty": 2.5, "unit_price": 10.0, "line_total": 25.0,
             "review": {}, "notes": {}},
        ], totals={}))
        self.assertEqual(sheet.cell(2, self.column("الكمية")).number_format, "#,##0.###")

    def test_currency_falls_back_to_a_plain_number_format(self):
        sheet = self.build(document(currency=""))
        self.assertEqual(sheet.cell(2, self.column("الإجمالي")).number_format, "#,##0.00")

    def test_a_hostile_currency_string_cannot_break_the_format(self):
        # The currency comes from a model reading pixels; it is untrusted input.
        fmt = excel_builder.money_format('X";@;[Red]')
        self.assertEqual(fmt.count('"'), 2)

    def test_column_widths_are_set_and_clamped(self):
        sheet = self.build()
        widths = [dimension.width for dimension in sheet.column_dimensions.values()]
        self.assertTrue(widths)
        for width in widths:
            self.assertGreaterEqual(width, excel_builder.MIN_WIDTH)
            self.assertLessEqual(width, excel_builder.MAX_WIDTH)

    def test_an_arabic_cell_reads_right_to_left(self):
        # Per cell, not per sheet: these tables mix an Arabic description with
        # an English SKU, and one sheet-wide setting misplaces the punctuation
        # of whichever language loses.
        sheet = self.build()
        cell = sheet.cell(2, self.column("الوصف"))
        self.assertEqual(cell.alignment.readingOrder, 2)
        self.assertEqual(cell.alignment.horizontal, "right")

    def test_an_english_cell_in_the_same_table_reads_left_to_right(self):
        sheet = self.build(document(items=[
            {"description": "Ladies' Garments", "qty": 2, "unit_price": 15.5,
             "line_total": 31.0, "review": {}, "notes": {}},
        ], totals={}))
        cell = sheet.cell(2, self.column("الوصف"))
        self.assertEqual(cell.alignment.readingOrder, 1)
        self.assertEqual(cell.alignment.horizontal, "left")

    def test_a_numeric_only_cell_is_left_to_the_context(self):
        self.assertEqual(excel_builder.reading_order("1024417"), 0)

    def test_an_empty_cell_is_left_to_the_context(self):
        self.assertEqual(excel_builder.reading_order(""), 0)

    def test_an_arabic_document_opens_right_to_left(self):
        sheet = self.build()
        self.assertTrue(sheet.sheet_view.rightToLeft)

    def test_an_english_document_stays_left_to_right(self):
        sheet = self.build(document(direction="ltr"))
        self.assertFalse(sheet.sheet_view.rightToLeft)

    def test_the_table_is_frozen_below_its_single_heading_row(self):
        sheet = self.build()
        self.assertEqual(sheet.freeze_panes, "A2")

    def test_the_heading_row_repeats_on_every_printed_page(self):
        sheet = self.build()
        self.assertEqual(str(sheet.print_title_rows).replace("$", ""), "1:1")


class ReviewTests(WorkbookCase):
    def test_a_flagged_cell_is_yellow_and_queued(self):
        """The queue points into the page sheet — the one a reviewer works in."""
        flagged = document()
        flagged["items"][1]["review"]["qty"] = True
        flagged["items"][1]["notes"]["qty"] = "راجع الكمية"
        self.build(flagged)
        qty = self.column("الكمية", self.document)
        second = self.item_row(2)
        self.assertIn("FFF2CC", str(self.document.cell(second, qty).fill.fgColor.rgb))
        _records, low, review_items, _template, _path = self.result
        self.assertEqual(low, 1)
        self.assertEqual(len(review_items), 1)
        self.assertEqual(review_items[0]["sheet"], self.document.title)
        self.assertEqual(review_items[0]["column"], qty)
        self.assertEqual(review_items[0]["row"], second)

    def test_a_flagged_formula_cell_queues_the_quantity_instead(self):
        # apply_review_file writes the corrected value straight into the cell,
        # so queueing the formula cell would replace the formula with a constant.
        flagged = document()
        flagged["items"][0]["review"]["line_total"] = True
        self.build(flagged)
        total = self.column("الإجمالي", self.document)
        first = self.item_row(1)
        self.assertTrue(str(self.document.cell(first, total).value).startswith("="))
        _records, _low, review_items, _template, _path = self.result
        self.assertEqual(len(review_items), 1)
        self.assertEqual(review_items[0]["column"], self.column("الكمية", self.document))

    def test_no_queued_cell_ever_points_at_a_formula(self):
        flagged = document()
        flagged["items"][0]["review"]["line_total"] = True
        flagged["items"][1]["review"]["unit_price"] = True
        flagged["totals_review"]["grand_total"] = True
        sheet = self.build(flagged)
        _records, _low, review_items, _template, _path = self.result
        for item in review_items:
            value = sheet.cell(item["row"], item["column"]).value
            self.assertFalse(
                isinstance(value, str) and value.startswith("="),
                f"عنصر مراجعة يشير إلى صيغة في {item['row']},{item['column']}",
            )

    def test_a_queued_cell_points_at_the_row_it_was_flagged_on(self):
        # The row is captured when the cell is written, not read off a variable
        # that has since moved on to the totals.
        flagged = document()
        flagged["items"][2]["review"]["qty"] = True
        self.build(flagged)
        _records, _low, review_items, _template, _path = self.result
        self.assertEqual(review_items[0]["row"], self.item_row(3))

    def test_a_flagged_total_is_highlighted(self):
        flagged = document()
        flagged["totals_review"]["subtotal"] = True
        flagged["totals_notes"]["subtotal"] = "المجموع لا يطابق"
        sheet = self.build(flagged)
        cell = sheet.cell(self.find_row("المجموع الفرعي"), self.column("الإجمالي"))
        self.assertIn("FFF2CC", str(cell.fill.fgColor.rgb))

    def test_the_review_queue_points_at_the_saved_workbook(self):
        flagged = document()
        flagged["items"][0]["review"]["qty"] = True
        self.build(flagged)
        _records, _low, review_items, _template, path = self.result
        self.assertEqual(review_items[0]["output"], str(path))


class StructureTests(WorkbookCase):
    def test_record_count_matches_the_item_count(self):
        self.build()
        self.assertEqual(self.result[0], 3)

    def test_one_sheet_per_document(self):
        # Two documents the merger kept apart stay two sheets.
        self.build(document(page=1), document(page=2, pages=[2]))
        self.assertEqual(len(self.pages), 2)
        self.assertEqual(self.result[0], 6)

    def test_the_workbook_opens_on_the_document_itself(self):
        """What the customer wants first is their invoice, not our notes on it.

        The workbook used to open on the review summary. It reads better as the
        document, the data, then the notes — in that order.
        """
        self.build()
        self.assertIs(self.book.worksheets[0], self.document)
        self.assertIs(self.book.worksheets[-1], self.summary)

    def test_an_english_document_is_written_in_english(self):
        """No Arabic anywhere in an English invoice's workbook.

        The complaint this guards against is a real one: an English invoice came
        back with Arabic headings over its figures, which stops the sheet being
        a transcription of the page it came from.
        """
        # Every scrap of content is English here, so anything Arabic left in the
        # workbook can only have come from the builder's own vocabulary.
        self.build(document(
            direction="ltr",
            title="TAX INVOICE",
            columns=["Item", "Qty", "Unit Price", "Amount"],
            column_roles=["description", "qty", "unit_price", "line_total"],
            items=[{"description": "Steel bracket", "qty": 2,
                    "unit_price": 15.5, "line_total": 31.0}],
            header={"supplier": "Northwind Ltd", "invoice_number": "INV-9"},
            totals={"subtotal": 31.0, "grand_total": 31.0},
        ))
        for sheet in self.book.worksheets:
            found = excel_builder.ARABIC.findall(self.text_of(sheet))
            self.assertFalse(found, f"Arabic in sheet {sheet.title!r}: {found[:8]}")
        self.assertFalse(self.sheet.sheet_view.rightToLeft)
        self.assertFalse(self.summary.sheet_view.rightToLeft)
        self.assertEqual(self.summary.title, "Review")

    def test_an_arabic_document_is_written_in_arabic_and_reads_right_to_left(self):
        self.build(document(direction="rtl"))
        self.assertTrue(self.sheet.sheet_view.rightToLeft)
        self.assertTrue(self.summary.sheet_view.rightToLeft)
        self.assertEqual(self.summary.title, "المراجعة")
        self.assertIn("ملخّص المراجعة", self.text_of(self.summary))

    def test_the_summary_lists_every_flagged_value_with_a_reason(self):
        self.build(document(items=[{
            "description": "قلم", "qty": 2, "unit_price": 15.5, "line_total": 34.0,
            "review": {"line_total": True}, "notes": {"line_total": "لا يطابق الضرب"},
        }], totals={}))
        self.assertIn("لا يطابق الضرب", self.text_of(self.summary))

    def test_the_summary_says_so_when_nothing_needs_review(self):
        self.build()
        self.assertIn("لا شيء", self.text_of(self.summary))

    def test_the_whole_table_can_be_sorted_and_filtered(self):
        sheet = self.build()
        # Including the header columns: filtering to one consignee is the first
        # thing anyone does with a merged manifest.
        self.assertEqual(sheet.auto_filter.ref, "A1:F4")

    def test_the_sheet_is_set_up_to_print_on_one_width(self):
        sheet = self.build()
        self.assertEqual(sheet.page_setup.fitToWidth, 1)
        self.assertTrue(sheet.sheet_properties.pageSetUpPr.fitToPage)

    def test_headings_follow_the_printed_column_order(self):
        # Item keys arrive in schema order; the printed table put الإجمالي first.
        self.build(document(
            columns=["الإجمالي", "سعر الوحدة", "الكمية", "الوصف"],
            column_roles=["line_total", "unit_price", "qty", "description"],
        ))
        self.assertEqual(self.column("الإجمالي"), 3)
        self.assertEqual(self.column("الوصف"), 6)

    def test_the_formula_follows_the_reordered_columns(self):
        from openpyxl.utils import get_column_letter

        sheet = self.build(document(
            columns=["الإجمالي", "سعر الوحدة", "الكمية", "الوصف"],
            column_roles=["line_total", "unit_price", "qty", "description"],
        ))
        qty = get_column_letter(self.column("الكمية"))
        price = get_column_letter(self.column("سعر الوحدة"))
        self.assertEqual(sheet.cell(2, self.column("الإجمالي")).value, f"={qty}2*{price}2")

    def test_a_recognised_column_is_not_promoted_to_the_front(self):
        """The sheet reads like the paper, not like the schema.

        An inventory sheet prints "Inward Quantity" and "Outward Quantity"
        before the "Quantity In Stock" that the formulas actually use. Pulling
        the recognised columns forward rearranged the customer's own table.
        """
        self.build(document(
            columns=["Product ID", "Product Name", "Inward Qty", "Qty In Stock", "Rate", "Amount"],
            column_roles=["sku", "description", "other", "qty", "unit_price", "line_total"],
            items=[{"sku": "MSG001", "description": "Item 1", "Inward Qty": "150",
                    "qty": 25, "unit_price": 1500.0, "line_total": 37500.0,
                    "review": {}, "notes": {}}],
            totals={},
        ))
        self.assertEqual(
            [cell.value for cell in self.sheet[1]][-6:],
            ["Product ID", "Product Name", "Inward Qty", "Qty In Stock", "Rate", "Amount"],
        )

    def test_an_unknown_field_becomes_a_plain_column(self):
        sheet = self.build(document(items=[
            {"description": "قلم", "qty": 2, "unit_price": 15.5, "line_total": 31.0,
             "warranty": "سنتان", "review": {}, "notes": {}},
        ], totals={}))
        self.assertEqual(sheet.cell(2, self.column("warranty")).value, "سنتان")


if __name__ == "__main__":
    unittest.main()
