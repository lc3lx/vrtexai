import unittest
from pathlib import Path
import sys

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from clean import correct_identifier, find_header_row, normalize_date, normalize_number, validate_detailed
from invoice import parse_invoice_table, totals_mismatch


TESTDATA = ROOT / "testdata"


class IdentifierTests(unittest.TestCase):
    def test_numeric_code_o_to_zero(self):
        self.assertEqual(correct_identifier("O46620", "ID_Code"), "046620")
        self.assertEqual(correct_identifier("09461O", "id_code"), "094610")

    def test_prefixed_segments(self):
        self.assertEqual(correct_identifier("CMP-95-O0", "رمز الشركة (ID)"), "CMP-95-00")
        self.assertEqual(correct_identifier("CMP-74-Il1", "id"), "CMP-74-111")
        self.assertEqual(correct_identifier("SHP-O5119-l", "رقم الشحنة"), "SHP-05119-1")

    def test_weight_units_not_corrupted(self):
        from clean import clean_text
        self.assertEqual(clean_text("197.67 tons", "وزن الحمل (طن)"), "197.67 tons")
        self.assertEqual(clean_text("565.03 طن", "وزن الحمل (طن)"), "565.03 طن")


class HeaderDetectionTests(unittest.TestCase):
    def test_complex_report_skips_title_rows(self):
        from openpyxl import load_workbook
        path = TESTDATA / "Complex_Table_Report.xlsx"
        self.assertTrue(path.is_file(), "missing testdata workbook")
        book = load_workbook(path, read_only=True, data_only=True)
        rows = list(book.active.iter_rows(values_only=True))
        book.close()
        header_at = find_header_row(rows)
        self.assertEqual(header_at, 5)
        headers = [str(value or "") for value in rows[header_at]]
        self.assertIn("رمز الشركة", headers[1])

    def test_simple_sheet_uses_first_row(self):
        from openpyxl import load_workbook
        path = TESTDATA / "Comprehensive_OCR_Test_Data.xlsx"
        self.assertTrue(path.is_file(), "missing testdata workbook")
        book = load_workbook(path, read_only=True, data_only=True)
        rows = list(book.active.iter_rows(values_only=True))
        book.close()
        self.assertEqual(find_header_row(rows), 0)


class CleaningTests(unittest.TestCase):
    def test_dates(self):
        self.assertEqual(normalize_date("17/09/2015"), "2015-09-17")
        self.assertEqual(normalize_date("2021-01-01"), "2021-01-01")

    def test_thousands_separators(self):
        self.assertEqual(normalize_number("6003,265,791"), "6003265791")
        self.assertEqual(normalize_number("1,234.50"), "1234.50")


class InvoiceTests(unittest.TestCase):
    def test_qty_price_mismatch_is_flagged(self):
        self.assertTrue(totals_mismatch("2", "10", "15"))
        self.assertFalse(totals_mismatch("2", "10", "20"))

    def test_header_separated_from_items(self):
        rows = [
            ["فاتورة ضريبية", "", "", ""],
            ["العميل: شركة الاختبار", "رقم الفاتورة: INV-9", "التاريخ: 01/02/2024", ""],
            ["الوصف", "الكمية", "سعر الوحدة", "المبلغ"],
            ["خدمة أ", "2", "10.00", "20.00"],
            ["خدمة ب", "1", "5.00", "5.00"],
            ["المجموع", "", "", "25.00"],
        ]
        parsed = parse_invoice_table(rows)
        self.assertEqual(len(parsed["items"]), 2)
        self.assertEqual(parsed["items"][0]["qty"], "2")
        self.assertTrue(parsed["header"].get("invoice_number") or parsed["header"].get("client_name"))


class MasterDataTests(unittest.TestCase):
    def test_exact_match_unchanged(self):
        result = validate_detailed("Acme", "supplier", {"supplier": ["Acme", "Globex"]})
        self.assertFalse(result.changed)
        self.assertEqual(result.value, "Acme")

    def test_fuzzy_match_corrects(self):
        result = validate_detailed("Acmee", "supplier", {"supplier": ["Acme"]})
        self.assertTrue(result.changed)
        self.assertEqual(result.value, "Acme")
        self.assertEqual(result.reason, "master-data")

    def test_ambiguous_match_not_auto_corrected(self):
        result = validate_detailed("Alpha Supplie", "supplier", {"supplier": ["Alpha Supplies", "Alpha Supplier"]})
        self.assertTrue(result.ambiguous)
        self.assertTrue(result.review)
        self.assertEqual(result.value, "Alpha Supplie")
        self.assertFalse(result.changed)

    def test_no_match_keeps_value(self):
        result = validate_detailed("Unknown Co", "supplier", {"supplier": ["Acme"]})
        self.assertEqual(result.value, "Unknown Co")
        self.assertFalse(result.review)


if __name__ == "__main__":
    unittest.main()
