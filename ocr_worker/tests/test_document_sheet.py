"""Tests for reproducing the page in Excel.

The customer's test is the one these follow: put the invoice beside the
workbook and they should read the same. Everything on the paper present,
nothing invented, in the order it was printed, each piece of data in its own
cell.

The fixture is ``data/7.png`` — a UAE tax invoice, right to left, with a company
heading, a tax number, two boxes of details side by side, a ten-line goods table
with its own footing row, a summary box, an amount in words, a note and a
signature line. The workbook it used to produce carried none of that: twelve
columns of mangled header fragments repeated down ten rows, then the goods, and
nothing else at all.
"""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import ai_extract
import document_sheet
import excel_builder
import paddle_vl
import table_probe

INVOICE_7 = """<p>مشرق التجار ذ م م</p>
<p>GHI، شارع ديف، ABC مبنى، 123</p>
<p>TRN: AE12DE123456</p>
<p>فاتورة الضرائب</p>
<p>رقم الفاتورة: ١    تاريخ: 12/27/2021</p>
<table>
<tr><td>معرف العميل: ش ب ط 001</td><td>نوع: Credit</td></tr>
<tr><td>اسم: أجهزة كمبيوتر الأسمنت</td><td>أيام: 15</td></tr>
<tr><td>عنوان: شارع ديف 1 - ABC</td><td>تاريخ الاستحقاق: 1/11/2022</td></tr>
</table>
<table>
<tr><td>المسلسل</td><td>معرف المنتج</td><td>وصف المنتج</td><td>معدل</td><td>الكمية</td><td>القيمة الضريبية</td><td>ضريبة %</td><td>مبلغ</td></tr>
<tr><td>1</td><td>ب ط 001</td><td>دبل إنسبيرون 1050</td><td>AED 25000</td><td>2</td><td>AED 50000</td><td>5%</td><td>AED 52500</td></tr>
<tr><td>2</td><td>ب ط 002</td><td>لينوفو 1-5125</td><td>AED 28000</td><td>2</td><td>AED 56000</td><td>5%</td><td>AED 58800</td></tr>
<tr><td>3</td><td>ب ط 009</td><td>لوحة الماوس</td><td>AED 100</td><td>2</td><td>AED 200</td><td>5%</td><td>AED 210</td></tr>
<tr><td colspan="4">مجموع</td><td>6</td><td>AED 106200</td><td></td><td>AED 111510</td></tr>
</table>
<p>المبلغ بالكلمات</p>
<p>مائة وأحد عشر ألف وخمسمائة وعشرة درهم</p>
<table>
<tr><td>ملخص</td><td></td></tr>
<tr><td>المبلغ خاضع للضريبة</td><td>AED 106200</td></tr>
<tr><td>مبلغ ضريبة القيمة المضافة</td><td>AED 5310</td></tr>
<tr><td>مبلغ الفاتورة</td><td>AED 111510</td></tr>
</table>
<p>الموقع لمعتمد    ختم الشركة</p>
<p>شكراً لك على التعامل معنا</p>"""


def build(html: str = INVOICE_7):
    """The whole path, from transcribed page to opened workbook."""
    from openpyxl import load_workbook

    page = {"result": table_probe._blocks_from_html(html), "markdown": ""}
    payload = paddle_vl.to_payload(page)
    document, blocking, _advisory = ai_extract.validate(payload, set())
    document["page"] = 1
    documents = ai_extract.merge_pages([document])

    directory = tempfile.TemporaryDirectory()
    destination = Path(directory.name) / "7.xlsx"
    result = excel_builder.write_ai_workbook(destination, Path("7.png"), documents)
    book = load_workbook(result[4])
    book.saved_path = result[4]
    return payload, document, blocking, book, directory


class PageFidelityTests(unittest.TestCase):
    """Everything on the paper reaches the sheet, in the order it was printed."""

    @classmethod
    def setUpClass(cls):
        cls.payload, cls.document, cls.blocking, cls.book, cls._dir = build()
        cls.sheet = cls.book.worksheets[0]
        cls.text = "\n".join(
            str(cell.value) for row in cls.sheet.iter_rows() for cell in row
            if isinstance(cell.value, str)
        )

    @classmethod
    def tearDownClass(cls):
        cls.book.close()
        cls._dir.cleanup()

    def row_of(self, text: str) -> int:
        for row in self.sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == text:
                    return cell.row
        raise AssertionError(f"«{text}» ليست في الورقة")

    def test_the_workbook_opens_on_the_page_itself(self):
        self.assertIs(self.book.worksheets[0], self.sheet)
        self.assertEqual(self.book.worksheets[-1].title, "المراجعة")

    def test_every_block_of_the_page_is_present(self):
        for printed in (
            "مشرق التجار ذ م م",              # the company
            "GHI، شارع ديف، ABC مبنى، 123",   # its address
            "AE12DE123456",                   # the tax number
            "فاتورة الضرائب",                  # the document's own title
            "معرف العميل: ش ب ط 001",          # the customer box
            "تاريخ الاستحقاق: 1/11/2022",      # the payment box
            "دبل إنسبيرون 1050",               # the goods
            "مجموع",                           # the table's footing row
            "المبلغ بالكلمات",                  # the amount in words
            "مبلغ الفاتورة",                    # the summary box
            "ختم الشركة",                      # the signature line
            "شكراً لك على التعامل معنا",        # the footer
        ):
            self.assertIn(printed, self.text, f"«{printed}» لم تصل إلى الورقة")

    def test_the_blocks_are_in_the_order_the_page_prints_them(self):
        order = [
            "مشرق التجار ذ م م", "فاتورة الضرائب", "معرف العميل: ش ب ط 001",
            "المسلسل", "مجموع", "المبلغ بالكلمات", "ملخص", "شكراً لك على التعامل معنا",
        ]
        rows = [self.row_of(text) for text in order]
        self.assertEqual(rows, sorted(rows), "ترتيب الأقسام لا يطابق ترتيب الورقة")

    def test_the_page_reads_right_to_left(self):
        self.assertTrue(self.sheet.sheet_view.rightToLeft)

    def test_the_masthead_is_given_the_weight_the_page_gives_it(self):
        # A transcribing model labels every paragraph "text", so nothing arrives
        # marked as the title; the first short line is where a document puts its
        # name, and it should not look like another field.
        heading = self.sheet.cell(self.row_of("مشرق التجار ذ م م"), 1)
        self.assertTrue(heading.font.bold)
        self.assertGreater(heading.font.size, 11)
        self.assertEqual(heading.alignment.horizontal, "center")

    def test_two_fields_on_one_printed_line_become_two_pairs_of_cells(self):
        # "رقم الفاتورة: ١    تاريخ: 12/27/2021" is four pieces of information.
        row = self.row_of("رقم الفاتورة")
        self.assertEqual(
            [self.sheet.cell(row, column).value for column in range(1, 5)],
            ["رقم الفاتورة", "١", "تاريخ", "12/27/2021"],
        )

    def test_a_box_of_details_keeps_one_row_per_line(self):
        # Its three rows used to be folded into a single row of column headings.
        first = self.row_of("معرف العميل: ش ب ط 001")
        self.assertEqual(self.sheet.cell(first, 2).value, "نوع: Credit")
        self.assertEqual(self.sheet.cell(first + 1, 1).value, "اسم: أجهزة كمبيوتر الأسمنت")
        self.assertEqual(self.sheet.cell(first + 2, 2).value, "تاريخ الاستحقاق: 1/11/2022")

    def test_the_goods_table_keeps_the_printed_column_order(self):
        row = self.row_of("المسلسل")
        self.assertEqual(
            [self.sheet.cell(row, column).value for column in range(1, 9)],
            ["المسلسل", "معرف المنتج", "وصف المنتج", "معدل", "الكمية",
             "القيمة الضريبية", "ضريبة %", "مبلغ"],
        )

    def test_the_footing_row_of_the_table_is_written_under_it(self):
        row = self.row_of("مجموع")
        self.assertEqual(self.sheet.cell(row, 5).value, 6)
        self.assertEqual(self.sheet.cell(row, 6).value, 106200)


class ValueTests(unittest.TestCase):
    """What is in the cells, once the page has been reproduced."""

    @classmethod
    def setUpClass(cls):
        cls.payload, cls.document, cls.blocking, cls.book, cls._dir = build()
        cls.sheet = cls.book.worksheets[0]

    @classmethod
    def tearDownClass(cls):
        cls.book.close()
        cls._dir.cleanup()

    def test_the_totals_are_read_from_the_summary_box(self):
        self.assertEqual(self.document["totals"]["subtotal"], 106200.0)
        self.assertEqual(self.document["totals"]["tax_amount"], 5310.0)
        self.assertEqual(self.document["totals"]["grand_total"], 111510.0)

    def test_the_footing_row_is_not_an_eleventh_product(self):
        self.assertEqual(len(self.document["items"]), 3)
        self.assertNotIn("مجموع", [item.get("description") for item in self.document["items"]])

    def test_money_printed_with_its_currency_is_still_a_number(self):
        # "AED 25000" measured as text, so no column could be the price and the
        # workbook came back with neither prices nor totals.
        self.assertEqual(self.document["items"][0]["unit_price"], 25000.0)
        self.assertEqual(self.document["items"][0]["qty"], 2.0)

    def test_the_columns_are_understood(self):
        self.assertEqual(
            self.payload["column_roles"],
            ["other", "other", "description", "unit_price", "qty", "line_total", "tax", "other"],
        )

    def test_the_arithmetic_holds_so_nothing_is_flagged(self):
        self.assertEqual(self.blocking, [])

    def test_a_product_code_stays_a_product_code(self):
        # "ب ط 001" used to reach the sheet as the number 1.
        row = next(
            cell.row for cell in self.sheet["C"]
            if isinstance(cell.value, str) and cell.value == "دبل إنسبيرون 1050"
        )
        self.assertEqual(self.sheet.cell(row, 2).value, "ب ط 001")

    def test_a_rate_is_stored_as_a_rate_and_shown_as_one(self):
        row = next(
            cell.row for cell in self.sheet["C"]
            if isinstance(cell.value, str) and cell.value == "دبل إنسبيرون 1050"
        )
        cell = self.sheet.cell(row, 7)
        self.assertEqual(cell.value, 0.05)
        self.assertIn("%", cell.number_format)

    def test_every_formula_shows_its_answer_without_recalculating(self):
        """What a viewer in Protected View sees.

        That is where the customer found an empty Amount column, an empty
        Subtotal and an empty Total: nothing recalculates there, so a formula
        with no stored result shows nothing at all. ``data_only=True`` reads the
        file that same way.
        """
        from openpyxl import load_workbook

        stored = load_workbook(self.book.saved_path, data_only=True)
        self.addCleanup(stored.close)
        checked = 0
        for sheet in self.book.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        seen = stored[sheet.title][cell.coordinate].value
                        self.assertIsNotNone(
                            seen, f"{sheet.title}!{cell.coordinate} would show empty"
                        )
                        checked += 1
        self.assertGreater(checked, 0, "no formula was written at all")

    def test_the_line_total_is_a_live_formula(self):
        row = next(
            cell.row for cell in self.sheet["C"]
            if isinstance(cell.value, str) and cell.value == "دبل إنسبيرون 1050"
        )
        self.assertTrue(str(self.sheet.cell(row, 6).value).startswith("="))


class DataSheetTests(unittest.TestCase):
    def test_the_flat_table_is_kept_as_a_second_sheet(self):
        # The page for reading; the flat table for sorting, filtering, pivoting.
        _payload, _document, _blocking, book, directory = build()
        self.addCleanup(directory.cleanup)
        self.addCleanup(book.close)
        self.assertEqual(book.sheetnames[1], "البيانات")
        self.assertEqual(book.sheetnames[-1], "المراجعة")


MANIFEST = """<p>inquire@impactrun.mail | Lincoln, NE 68501 | 222 555 7777 | Template.net</p>
<p>Shipping Manifest</p>
<p>Recipient Information</p>
<table>
<tr><td>Recipient Name</td><td>KinFinity</td></tr>
<tr><td>Recipient Email</td><td>inquire@kinfinity.mail</td></tr>
<tr><td>Recipient Address</td><td>Greensboro, NC 27401</td></tr>
<tr><td>Recipient Phone</td><td>222 555 7777</td></tr>
</table>
<p>Shipping Details</p>
<table>
<tr><td>Item Description</td><td>Quantity</td><td>Weight</td><td>Dimensions</td><td>Total Value</td></tr>
<tr><td>Apple MacBook Pro 16"</td><td>10</td><td>2 kg</td><td>36x24x2cm</td><td>$25,000</td></tr>
<tr><td>Logitech MX Master 3</td><td>5</td><td>0.2 kg</td><td>12x8x4cm</td><td>$500</td></tr>
<tr><td>Bose QuietComfort 35 II</td><td>3</td><td>0.5 kg</td><td>18x18x8cm</td><td>$600</td></tr>
</table>"""


class ManifestTests(unittest.TestCase):
    """A shipping manifest in English, and the four faults it exposed."""

    @classmethod
    def setUpClass(cls):
        cls.payload, cls.document, _blocking, cls.book, cls._dir = build(MANIFEST)
        cls.sheet = cls.book.worksheets[0]

    @classmethod
    def tearDownClass(cls):
        cls.book.close()
        cls._dir.cleanup()

    def row_of(self, text: str) -> int:
        for row in self.sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip() == text:
                    return cell.row
        raise AssertionError(f"'{text}' is not on the sheet")

    def test_the_recipient_box_keeps_one_row_per_field(self):
        # All four rows arrived crammed into one: three labels in one cell and
        # three values in another.
        first = self.row_of("Recipient Name")
        for offset, (label, value) in enumerate((
            ("Recipient Name", "KinFinity"),
            ("Recipient Email", "inquire@kinfinity.mail"),
            ("Recipient Address", "Greensboro, NC 27401"),
            ("Recipient Phone", "222 555 7777"),
        )):
            self.assertEqual(self.sheet.cell(first + offset, 1).value, label)
            self.assertEqual(self.sheet.cell(first + offset, 2).value, value)

    def test_the_phone_number_is_still_a_phone_number(self):
        # It reached the sheet as the number 2,225,557,777.
        cell = self.sheet.cell(self.row_of("Recipient Phone"), 2)
        self.assertEqual(cell.value, "222 555 7777")
        self.assertNotIsInstance(cell.value, (int, float))

    def test_the_amounts_carry_the_dollar_sign_the_page_prints(self):
        self.assertEqual(self.document["currency"], "USD")
        row = self.row_of("Apple MacBook Pro 16\"")
        cell = self.sheet.cell(row, 5)
        self.assertEqual(cell.value, 25000)
        self.assertIn("$", cell.number_format)

    def test_a_whole_quantity_shows_without_a_trailing_point(self):
        cell = self.sheet.cell(self.row_of("Apple MacBook Pro 16\""), 2)
        self.assertEqual(cell.number_format, "#,##0")

    def test_a_page_that_prints_no_total_still_gets_one(self):
        # The manifest has a column of amounts and no sum of them. The sheet
        # adds them up, as a formula, under a label that says it calculated it.
        row = self.row_of("Total (calculated)")
        formula = self.sheet.cell(row, 5).value
        self.assertTrue(str(formula).startswith("=SUM("))
        self.assertIn("$", self.sheet.cell(row, 5).number_format)

    def test_the_calculated_total_shows_its_figure(self):
        # It came back blank: a formula with no stored result shows nothing
        # wherever nothing recalculates, which is where the customer opened it.
        from openpyxl import load_workbook

        stored = load_workbook(self.book.saved_path, data_only=True)
        self.addCleanup(stored.close)
        row = self.row_of("Total (calculated)")
        self.assertEqual(stored[self.sheet.title].cell(row, 5).value, 26100)


PRACTICE = """<p>Excel Data Entry Practice Exercises PDF</p>
<table>
<tr><td></td><td>Date</td><td>Item</td><td>Price</td><td>Unit</td><td>Amount</td><td>Discount</td><td>Net Amount</td><td>Sales Tax</td><td>Total</td></tr>
<tr><td></td><td>28-Oct-2022</td><td>Pencil</td><td>$0.27</td><td>26</td><td>$7.02</td><td>$0.21</td><td>$6.81</td><td>$0.68</td><td>$7.49</td></tr>
<tr><td></td><td>28-Oct-2022</td><td>Gel Pen</td><td>$1.40</td><td>6</td><td>$8.40</td><td>$0.42</td><td>$7.98</td><td>$0.80</td><td>$8.78</td></tr>
<tr><td></td><td>28-Oct-2022</td><td>Calculator</td><td>$7.39</td><td>30</td><td>$221.70</td><td>$11.09</td><td>$210.62</td><td>$21.06</td><td>$231.68</td></tr>
</table>"""


class PracticeSheetTests(unittest.TestCase):
    """A nine-column ledger, and what a phone viewer made of it."""

    @classmethod
    def setUpClass(cls):
        cls.payload, cls.document, _blocking, cls.book, cls._dir = build(PRACTICE)
        cls.sheet = cls.book.worksheets[0]

    @classmethod
    def tearDownClass(cls):
        cls.book.close()
        cls._dir.cleanup()

    def heading_row(self) -> int:
        for row in self.sheet.iter_rows():
            for cell in row:
                if cell.value == "Date":
                    return cell.row
        raise AssertionError("the table's heading row is missing")

    def test_the_margin_column_does_not_become_a_column(self):
        # It shifted every heading one place right of its own data.
        row = self.heading_row()
        self.assertEqual(self.sheet.cell(row, 1).value, "Date")
        self.assertEqual(
            [self.sheet.cell(row, n).value for n in range(1, 10)],
            ["Date", "Item", "Price", "Unit", "Amount", "Discount",
             "Net Amount", "Sales Tax", "Total"],
        )

    def test_every_money_column_carries_the_currency(self):
        # "Net Amount" and "Total" have no role the resolver knows, and came out
        # as bare numbers beside four columns that showed dollars.
        first = self.heading_row() + 1
        for column in (3, 5, 6, 7, 8, 9):
            self.assertIn(
                "$", self.sheet.cell(first, column).number_format,
                f"column {column} lost its currency",
            )

    def test_no_column_is_formatted_with_a_hash_placeholder(self):
        first = self.heading_row() + 1
        for column in range(1, 10):
            decimals = self.sheet.cell(first, column).number_format.partition(".")[2]
            self.assertNotIn("#", decimals)


class SplitLineTests(unittest.TestCase):
    def test_a_line_of_two_labelled_fields_becomes_four_cells(self):
        self.assertEqual(
            document_sheet.split_line("رقم الفاتورة: ١    تاريخ: 12/27/2021"),
            [("رقم الفاتورة", True), ("١", False), ("تاريخ", True), ("12/27/2021", False)],
        )

    def test_a_clock_is_not_a_label(self):
        self.assertEqual(
            document_sheet.split_line("Time 3:00 PM"), [("Time 3:00 PM", False)]
        )

    def test_a_plain_sentence_stays_one_cell(self):
        self.assertEqual(
            document_sheet.split_line("شكراً لك على التعامل معنا"),
            [("شكراً لك على التعامل معنا", False)],
        )

    def test_two_unlabelled_fields_are_split_by_the_gap_between_them(self):
        self.assertEqual(
            document_sheet.split_line("الموقع لمعتمد    ختم الشركة"),
            [("الموقع لمعتمد", False), ("ختم الشركة", False)],
        )


if __name__ == "__main__":
    unittest.main()
