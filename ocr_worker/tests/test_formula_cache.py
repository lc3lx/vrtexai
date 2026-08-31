"""Tests for the formula results stored beside the formulas.

The bug these guard against showed the customer an invoice with its Amount
column empty, its Subtotal empty and its Total empty. Every one of those cells
held a formula, and openpyxl writes a formula with no result — so anything that
will not recalculate shows the cell blank. Excel's Protected View, which is how
every downloaded file opens, is exactly such a thing.

``data_only=True`` reads the stored result rather than the formula, which makes
it the same view of the file that Protected View and every preview pane has. It
is the right assertion here for that reason: if these pass, the numbers are
visible without anything having to compute them.
"""
from __future__ import annotations

import sys
import tempfile
import unittest
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import formula_cache


def workbook(cells: dict[str, object], title: str = "Sheet") -> Path:
    """A saved workbook whose formulas have no results, as openpyxl leaves them."""
    from openpyxl import Workbook

    book = Workbook()
    sheet = book.active
    sheet.title = title
    for coordinate, value in cells.items():
        sheet[coordinate] = value
    directory = tempfile.mkdtemp()
    path = Path(directory) / "book.xlsx"
    book.save(path)
    book.close()
    return path


def stored(path: Path, coordinate: str, title: str = "Sheet"):
    """What a reader that does not recalculate sees in a cell."""
    from openpyxl import load_workbook

    book = load_workbook(path, data_only=True)
    try:
        return book[title][coordinate].value
    finally:
        book.close()


class CachedValueTests(unittest.TestCase):
    def test_a_formula_shows_nothing_until_its_result_is_stored(self):
        # The defect itself, reproduced: this is what the customer opened.
        path = workbook({"A1": 2, "B1": 3, "C1": "=A1*B1"})
        self.assertIsNone(stored(path, "C1"))

    def test_storing_the_result_makes_the_cell_readable(self):
        path = workbook({"A1": 2, "B1": 3, "C1": "=A1*B1"})
        self.assertEqual(formula_cache.cache_formula_values(path, {"Sheet": {"C1": 6}}), 1)
        self.assertEqual(stored(path, "C1"), 6)

    def test_the_formula_itself_survives(self):
        # The point is a cell that both shows its answer and stays live.
        from openpyxl import load_workbook

        path = workbook({"A1": 2, "B1": 3, "C1": "=A1*B1"})
        formula_cache.cache_formula_values(path, {"Sheet": {"C1": 6}})
        book = load_workbook(path)
        self.addCleanup(book.close)
        self.assertEqual(book["Sheet"]["C1"].value, "=A1*B1")

    def test_every_shape_of_formula_the_builder_writes(self):
        path = workbook({
            "A1": 2, "B1": 15.5, "C1": "=A1*B1",
            "C2": "=SUM(C1:C1)", "C3": "=C2*0.15", "C4": "=C2+C3",
        })
        filled = formula_cache.cache_formula_values(path, {"Sheet": {
            "C1": 31.0, "C2": 31.0, "C3": 4.65, "C4": 35.65,
        }})
        self.assertEqual(filled, 4)
        self.assertEqual(stored(path, "C4"), 35.65)

    def test_a_sheet_named_in_arabic_is_found(self):
        # Sheet names reach the file as UTF-8 in the workbook's own XML.
        path = workbook({"A1": 1, "B1": "=A1*2"}, title="المراجعة")
        formula_cache.cache_formula_values(path, {"المراجعة": {"B1": 2}})
        self.assertEqual(stored(path, "B1", "المراجعة"), 2)

    def test_a_cell_no_value_was_given_for_is_left_alone(self):
        path = workbook({"A1": 2, "B1": "=A1*2", "C1": "=A1*3"})
        formula_cache.cache_formula_values(path, {"Sheet": {"B1": 4}})
        self.assertEqual(stored(path, "B1"), 4)
        self.assertIsNone(stored(path, "C1"))

    def test_a_value_that_is_not_a_number_is_refused(self):
        path = workbook({"A1": 2, "B1": "=A1*2"})
        self.assertEqual(
            formula_cache.cache_formula_values(path, {"Sheet": {"B1": "four"}}), 0
        )
        self.assertIsNone(stored(path, "B1"))

    def test_nothing_to_do_leaves_the_file_untouched(self):
        path = workbook({"A1": 1})
        before = path.read_bytes()
        self.assertEqual(formula_cache.cache_formula_values(path, {}), 0)
        self.assertEqual(path.read_bytes(), before)

    def test_the_workbook_still_opens_after_rewriting(self):
        # The file is rebuilt as a zip; a corrupt one would be far worse than a
        # blank cell, so the whole book is reopened and read back.
        from openpyxl import load_workbook

        path = workbook({"A1": 2, "B1": 3, "C1": "=A1*B1", "D1": "text"})
        formula_cache.cache_formula_values(path, {"Sheet": {"C1": 6}})
        book = load_workbook(path)
        self.addCleanup(book.close)
        self.assertEqual(book["Sheet"]["A1"].value, 2)
        self.assertEqual(book["Sheet"]["D1"].value, "text")

    def test_a_file_that_is_not_a_workbook_is_survivable(self):
        broken = Path(tempfile.mkdtemp()) / "not.xlsx"
        broken.write_bytes(b"not a zip at all")
        self.assertEqual(formula_cache.cache_formula_values(broken, {"Sheet": {"A1": 1}}), 0)
        self.assertEqual(broken.read_bytes(), b"not a zip at all")


if __name__ == "__main__":
    unittest.main()
